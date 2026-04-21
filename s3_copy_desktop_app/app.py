"""Tkinter desktop app for non-destructive S3-to-S3 copy operations on macOS."""

from __future__ import annotations

import csv
import gc
import hashlib
import json
import os
from pathlib import Path
import queue
import shutil
import subprocess
import sys
import threading
import time
import webbrowser
from dataclasses import asdict, dataclass, replace
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

if sys.platform.startswith("win"):
    import winsound

from .config_store import AppConfig, load_config, save_config
from .credentials_store import (
    AwsCredentials,
    KeychainOwnerConflictError,
    clear_credentials,
    load_credentials,
    save_credentials,
)
from .s3_service import (
    DestinationExistsError,
    S3ListedObject,
    S3ObjectRef,
    UserVisibleError,
    copy_object,
    create_s3_client,
    delete_object,
    list_objects_under_prefix,
    list_objects_with_metadata_under_prefix,
    object_exists,
    prefix_exists,
    upload_local_file,
)
from .validators import (
    ResolvedS3Paths,
    UserInput,
    build_paths,
    join_key_parts,
    sanitize_filename,
    sanitize_folder_path,
    sanitize_user_input,
    validate_paths_not_identical,
    validate_user_input,
)

APP_TITLE = os.getenv("S3_APP_TITLE", "s3Organizer")
APP_FILE_SLUG = os.getenv("S3_APP_FILE_SLUG", "s3Organizer")
SIMPLIFIED_BULK_REQUIRE_DRY_RUN = os.getenv("S3_SIMPLIFIED_BULK_REQUIRE_DRY_RUN", "1") != "0"
POWER_MODE = APP_FILE_SLUG.lower() == "powers3browser"
IS_WINDOWS = sys.platform.startswith("win")
IS_MACOS = sys.platform == "darwin"
WINDOW_WIDTH = 980 if POWER_MODE else 760
WINDOW_HEIGHT = 940
VOICE_COMMAND = ["say", "-v", "Samantha", "Copy Complete"]
DING_COMMAND = ["afplay", "/System/Library/Sounds/Glass.aiff"]
CURRENT_BLOCK_BG = "#d9edff"
CURRENT_FIELD_BG = "#f5fbff"
DESIRED_BLOCK_BG = "#9ecbf0"
DESIRED_FIELD_BG = "#eef7ff"
RESULTS_BLOCK_BG = "#6fa8da"
RESULTS_FIELD_BG = "#e3f0fb"
SECTION_TEXT_COLOR = "#0d2d4d"
CLEAR_BUTTON_WIDTH = 6
SIMPLIFIED_BULK_REQUIRED_COLUMNS = ("source_uri", "destination_uri")
BULK_FOLDER_REQUIRED_COLUMNS = ("source_folder_uri", "destination_folder_uri")
SIMPLIFIED_BULK_CHECKPOINT_DIR = Path.home() / "Library" / "Application Support" / APP_TITLE / "checkpoints"


def _configure_undo(widget: tk.Misc) -> None:
    """Enable undo support when the Tk widget exposes it."""
    try:
        widget.configure(undo=True, autoseparators=True, maxundo=-1)
    except tk.TclError:
        pass


def _credential_store_label() -> str:
    if IS_MACOS:
        return "macOS Keychain"
    if IS_WINDOWS:
        return "Windows Credential Manager"
    return "system credential store"


def _invoke_text_edit(widget: tk.Misc, operation: str) -> str | None:
    """Route undo/redo to the focused Tk text widget."""
    try:
        state = str(widget.cget("state"))
    except tk.TclError:
        state = "normal"

    if state in {"disabled", "readonly"}:
        return None

    try:
        widget.tk.call(widget._w, "edit", operation)
    except tk.TclError:
        return "break"
    return "break"


class EntryUndoManager:
    """Track entry field history so Cmd+Z works consistently on macOS."""

    def __init__(self) -> None:
        self._variables_by_name: dict[str, tk.StringVar] = {}
        self._undo_stacks: dict[str, list[str]] = {}
        self._redo_stacks: dict[str, list[str]] = {}
        self._last_values: dict[str, str] = {}
        self._restoring = False

    def register_variable(self, variable: tk.StringVar) -> None:
        variable_name = str(variable)
        if variable_name in self._variables_by_name:
            return

        self._variables_by_name[variable_name] = variable
        self._undo_stacks[variable_name] = []
        self._redo_stacks[variable_name] = []
        self._last_values[variable_name] = variable.get()
        variable.trace_add("write", lambda *_args, name=variable_name: self._on_variable_changed(name))

    def _on_variable_changed(self, variable_name: str) -> None:
        if self._restoring:
            return

        variable = self._variables_by_name[variable_name]
        current_value = variable.get()
        last_value = self._last_values[variable_name]
        if current_value == last_value:
            return

        self._undo_stacks[variable_name].append(last_value)
        if len(self._undo_stacks[variable_name]) > 100:
            self._undo_stacks[variable_name].pop(0)
        self._redo_stacks[variable_name].clear()
        self._last_values[variable_name] = current_value

    def _variable_name_from_widget(self, widget: tk.Misc) -> str | None:
        try:
            variable_name = str(widget.cget("textvariable"))
        except tk.TclError:
            return None

        if not variable_name:
            return None
        if variable_name not in self._variables_by_name:
            return None
        return variable_name

    def undo_from_widget(self, widget: tk.Misc) -> str | None:
        variable_name = self._variable_name_from_widget(widget)
        if not variable_name:
            return _invoke_text_edit(widget, "undo")

        undo_stack = self._undo_stacks[variable_name]
        if not undo_stack:
            return "break"

        variable = self._variables_by_name[variable_name]
        current_value = variable.get()
        previous_value = undo_stack.pop()
        self._redo_stacks[variable_name].append(current_value)
        self._restore_value(widget, variable_name, previous_value)
        return "break"

    def redo_from_widget(self, widget: tk.Misc) -> str | None:
        variable_name = self._variable_name_from_widget(widget)
        if not variable_name:
            return _invoke_text_edit(widget, "redo")

        redo_stack = self._redo_stacks[variable_name]
        if not redo_stack:
            return "break"

        variable = self._variables_by_name[variable_name]
        current_value = variable.get()
        next_value = redo_stack.pop()
        self._undo_stacks[variable_name].append(current_value)
        self._restore_value(widget, variable_name, next_value)
        return "break"

    def _restore_value(self, widget: tk.Misc, variable_name: str, value: str) -> None:
        self._restoring = True
        try:
            variable = self._variables_by_name[variable_name]
            variable.set(value)
            self._last_values[variable_name] = value
            try:
                widget.icursor("end")
            except tk.TclError:
                pass
        finally:
            self._restoring = False


@dataclass
class DirectUploadItem:
    label: str
    local_path: str
    destination_ref: S3ObjectRef
    destination_uri: str


@dataclass
class SimplifiedBulkCsvPreview:
    row_count: int
    first_source_uri: str
    first_destination_uri: str


@dataclass
class SimplifiedBulkCopyReportRow:
    row_label: str
    source_uri: str
    destination_uri: str
    status: str
    message: str
    destination_folder_uri: str = ""
    destination_folder_status: str = ""


@dataclass
class FolderCopyPreview:
    object_count: int
    first_source_uri: str
    first_destination_uri: str
    folder_pair_count: int = 1


@dataclass
class InventoryPreview:
    object_count: int
    first_object_uri: str


class DeferredOverwriteError(UserVisibleError):
    """Raised when overwrite handling is intentionally deferred to the end of a bulk run."""


class SettingsDialog(tk.Toplevel):
    """Hidden settings for fixed S3 routing and optional keychain credentials."""

    def __init__(
        self,
        parent: tk.Tk,
        config: AppConfig,
        session_credentials: AwsCredentials | None,
        use_session_only: bool,
        on_save,
    ) -> None:
        super().__init__(parent)
        self.title("Settings")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self._undo_manager = EntryUndoManager()

        self._on_save = on_save
        self._config = config
        self._session_credentials = session_credentials
        self._use_session_only = use_session_only

        self.columnconfigure(1, weight=1)

        padding = {"padx": 10, "pady": 6}

        ttk.Label(self, text="Source Bucket").grid(row=0, column=0, sticky="w", **padding)
        self.source_bucket_var = tk.StringVar(value=config.source_bucket)
        self._build_entry_row(row=0, variable=self.source_bucket_var)

        ttk.Label(self, text="Source Prefix").grid(row=1, column=0, sticky="w", **padding)
        self.source_prefix_var = tk.StringVar(value=config.source_prefix)
        self._build_entry_row(row=1, variable=self.source_prefix_var)

        source_example = (
            "Example from command:\n"
            "aws s3 cp s3://your-source-bucket/path/to/source-file.mov ...\n"
            "Source Bucket = your-source-bucket\n"
            "Source Prefix = path/to\n"
            "Leave Source Prefix blank if file is in bucket root."
        )
        ttk.Label(self, text=source_example, wraplength=560).grid(
            row=2, column=0, columnspan=3, sticky="w", padx=10, pady=(0, 8)
        )

        ttk.Label(self, text="Destination Bucket").grid(row=3, column=0, sticky="w", **padding)
        self.dest_bucket_var = tk.StringVar(value=config.dest_bucket)
        self._build_entry_row(row=3, variable=self.dest_bucket_var)

        ttk.Label(self, text="Destination Prefix").grid(row=4, column=0, sticky="w", **padding)
        self.dest_prefix_var = tk.StringVar(value=config.dest_prefix)
        self._build_entry_row(row=4, variable=self.dest_prefix_var)

        destination_example = (
            "Destination example from same command:\n"
            "aws s3 cp ... s3://your-destination-bucket/path/to/destination-file.mov\n"
            "Destination Bucket = your-destination-bucket\n"
            "Destination Prefix = path/to\n"
            "Leave Destination Prefix blank to write to bucket root."
        )
        ttk.Label(self, text=destination_example, wraplength=560).grid(
            row=5, column=0, columnspan=3, sticky="w", padx=10, pady=(0, 8)
        )

        ttk.Label(self, text="AWS Region (optional)").grid(row=6, column=0, sticky="w", **padding)
        self.aws_region_var = tk.StringVar(value=config.aws_region)
        self._build_entry_row(row=6, variable=self.aws_region_var)

        ttk.Separator(self, orient="horizontal").grid(
            row=7, column=0, columnspan=3, sticky="ew", padx=10, pady=(12, 8)
        )

        ttk.Label(self, text="AWS Access Key ID").grid(row=8, column=0, sticky="w", **padding)
        self.access_key_var = tk.StringVar()
        self.access_key_entry = self._build_entry_row(row=8, variable=self.access_key_var, show="*")

        ttk.Label(self, text="AWS Secret Access Key").grid(row=9, column=0, sticky="w", **padding)
        self.secret_key_var = tk.StringVar()
        self.secret_key_entry = self._build_entry_row(row=9, variable=self.secret_key_var, show="*")

        ttk.Label(self, text="AWS Session Token (optional)").grid(row=10, column=0, sticky="w", **padding)
        self.session_token_var = tk.StringVar()
        self.session_token_entry = self._build_entry_row(row=10, variable=self.session_token_var, show="*")

        self.show_credentials_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            self,
            text="Show credentials",
            variable=self.show_credentials_var,
            command=self._toggle_credential_visibility,
        ).grid(row=11, column=0, columnspan=3, sticky="w", padx=10, pady=(2, 2))

        self.session_only_var = tk.BooleanVar(value=use_session_only)
        ttk.Checkbutton(
            self,
            text="Use entered credentials for this app session only (do not save to Keychain)",
            variable=self.session_only_var,
        ).grid(row=12, column=0, columnspan=3, sticky="w", padx=10, pady=(2, 2))

        credential_note = (
            "Keychain mode: entered credentials are saved for reuse. "
            "Session-only mode: credentials are kept in memory and cleared when app closes."
        )
        ttk.Label(self, text=credential_note, wraplength=560).grid(
            row=13, column=0, columnspan=3, sticky="w", padx=10, pady=(2, 10)
        )

        button_frame = ttk.Frame(self)
        button_frame.grid(row=14, column=0, columnspan=3, sticky="e", padx=10, pady=(4, 10))

        ttk.Button(button_frame, text="Clear Stored Credentials", command=self._clear_stored_credentials).grid(
            row=0, column=0, padx=(0, 8)
        )
        ttk.Button(button_frame, text="Cancel", command=self.destroy).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(button_frame, text="Save", command=self._save).grid(row=0, column=2)

        self._load_stored_credentials()
        self._bind_standard_shortcuts()

    def _build_entry_row(self, row: int, variable: tk.StringVar, show: str | None = None) -> ttk.Entry:
        entry_kwargs = {"textvariable": variable, "width": 58}
        if show is not None:
            entry_kwargs["show"] = show
        entry = ttk.Entry(self, **entry_kwargs)
        self._undo_manager.register_variable(variable)
        entry.grid(row=row, column=1, sticky="ew", padx=10, pady=6)
        ttk.Button(
            self,
            text="Clear",
            width=CLEAR_BUTTON_WIDTH,
            command=lambda value=variable: value.set(""),
        ).grid(row=row, column=2, sticky="w", padx=(0, 10), pady=6)
        return entry

    def _bind_standard_shortcuts(self) -> None:
        for entry_class in ("Entry", "TEntry"):
            self.bind_class(entry_class, "<Command-z>", self._handle_undo_shortcut, add="+")
            self.bind_class(entry_class, "<Command-Z>", self._handle_redo_shortcut, add="+")
            self.bind_class(entry_class, "<Control-z>", self._handle_undo_shortcut, add="+")
            self.bind_class(entry_class, "<Control-Z>", self._handle_redo_shortcut, add="+")
        self.bind_class("Text", "<Command-z>", self._handle_undo_shortcut, add="+")
        self.bind_class("Text", "<Command-Z>", self._handle_redo_shortcut, add="+")
        self.bind_class("Text", "<Control-z>", self._handle_undo_shortcut, add="+")
        self.bind_class("Text", "<Control-Z>", self._handle_redo_shortcut, add="+")

    def _handle_undo_shortcut(self, event) -> str | None:
        return self._undo_manager.undo_from_widget(event.widget)

    def _handle_redo_shortcut(self, event) -> str | None:
        return self._undo_manager.redo_from_widget(event.widget)

    def _toggle_credential_visibility(self) -> None:
        show_character = "" if self.show_credentials_var.get() else "*"
        self.access_key_entry.configure(show=show_character)
        self.secret_key_entry.configure(show=show_character)
        self.session_token_entry.configure(show=show_character)

    def _load_stored_credentials(self) -> None:
        if self._session_credentials:
            self.access_key_var.set(self._session_credentials.access_key_id)
            self.secret_key_var.set(self._session_credentials.secret_access_key)
            self.session_token_var.set(self._session_credentials.session_token)
        return

    def _clear_stored_credentials(self) -> None:
        if not messagebox.askyesno(
            "Clear Credentials",
            f"Remove saved AWS credentials from {_credential_store_label()}?",
            parent=self,
        ):
            return

        try:
            clear_credentials()
        except RuntimeError as error:
            messagebox.showerror("Settings", str(error), parent=self)
            return

        self.access_key_var.set("")
        self.secret_key_var.set("")
        self.session_token_var.set("")
        messagebox.showinfo("Settings", "Stored credentials cleared.", parent=self)

    def _save(self) -> None:
        source_bucket = self.source_bucket_var.get().strip()
        source_prefix = sanitize_folder_path(self.source_prefix_var.get())
        dest_bucket = self.dest_bucket_var.get().strip()
        dest_prefix = sanitize_folder_path(self.dest_prefix_var.get())
        aws_region = self.aws_region_var.get().strip()

        if not source_bucket or not dest_bucket:
            messagebox.showerror(
                "Settings",
                "Source and Destination bucket are required. Prefixes can be blank.",
                parent=self,
            )
            return

        new_config = replace(
            self._config,
            source_bucket=source_bucket,
            source_prefix=source_prefix,
            dest_bucket=dest_bucket,
            dest_prefix=dest_prefix,
            aws_region=aws_region,
        )

        access_key = self.access_key_var.get().strip()
        secret_key = self.secret_key_var.get().strip()
        session_token = self.session_token_var.get().strip()

        if bool(access_key) ^ bool(secret_key):
            messagebox.showerror(
                "Settings",
                "Provide both AWS Access Key ID and Secret Access Key, or leave both blank.",
                parent=self,
            )
            return

        credential_mode = "session" if self.session_only_var.get() else "keychain"
        session_credentials: AwsCredentials | None = None
        entered_credentials: AwsCredentials | None = None
        if access_key and secret_key:
            entered_credentials = AwsCredentials(
                access_key_id=access_key,
                secret_access_key=secret_key,
                session_token=session_token,
            )

        try:
            if entered_credentials:
                if credential_mode == "session":
                    session_credentials = entered_credentials
                else:
                    save_credentials(entered_credentials)
        except KeychainOwnerConflictError:
            use_session_only = messagebox.askyesno(
                "Credential Store Blocked",
                (
                    f"{_credential_store_label()} blocked saving these credentials.\n\n"
                    "Switch to session-only mode and continue without saving them?"
                ),
                parent=self,
            )
            if not use_session_only:
                messagebox.showerror(
                    "Settings",
                    "Credentials were not saved. You can still use session-only mode.",
                    parent=self,
                )
                return

            credential_mode = "session"
            self.session_only_var.set(True)
            session_credentials = entered_credentials
            messagebox.showinfo(
                "Settings",
                "Session-only mode enabled. Credentials will be used for this app run only.",
                parent=self,
            )
        except RuntimeError as error:
            messagebox.showerror("Settings", str(error), parent=self)
            return

        try:
            # Persist keychain as the default startup mode.
            # Session-only mode remains a per-run option chosen in the UI.
            save_config(replace(new_config, credential_mode="keychain"))
        except RuntimeError as error:
            messagebox.showerror("Settings", str(error), parent=self)
            return

        new_config = replace(new_config, credential_mode=credential_mode)
        self._on_save(new_config, session_credentials)
        messagebox.showinfo("Settings", "Settings saved.", parent=self)
        self.destroy()


class BulkCopyDialog(tk.Toplevel):
    """Table-style input window for optional bulk copy jobs."""

    S3_COPY_COLUMNS = (
        ("title", "Title", 120),
        ("current_file_name", "Current File Name", 180),
        ("current_caption_name", "Current Caption Name (optional)", 210),
        ("desired_move_folder", "Desired Move Folder", 180),
        ("desired_name", "Desired Name", 180),
        ("desired_caption_name", "Desired Caption Name (optional)", 220),
    )
    DIRECT_UPLOAD_COLUMNS = (
        ("title", "Title #", 120),
        ("local_file_path", "Local File Name", 280),
        ("desired_move_folder", "Desired Move Folder", 180),
        ("desired_name", "Desired Name", 180),
    )
    S3_COPY_HEADER_ALIASES = {
        "title": "title",
        "current file name": "current_file_name",
        "current filename": "current_file_name",
        "current_caption_name": "current_caption_name",
        "current caption name": "current_caption_name",
        "desired move folder": "desired_move_folder",
        "desired_move_folder": "desired_move_folder",
        "desired name": "desired_name",
        "desired_name": "desired_name",
        "desired caption name": "desired_caption_name",
        "desired_caption_name": "desired_caption_name",
    }
    DIRECT_UPLOAD_HEADER_ALIASES = {
        "title": "title",
        "title #": "title",
        "title number": "title",
        "local file name": "local_file_path",
        "local filename": "local_file_path",
        "local file path": "local_file_path",
        "local filepath": "local_file_path",
        "local file": "local_file_path",
        "local_file_path": "local_file_path",
        "desired move folder": "desired_move_folder",
        "desired_move_folder": "desired_move_folder",
        "desired name": "desired_name",
        "desired_name": "desired_name",
    }

    def __init__(self, parent: tk.Tk, mode: str, on_run) -> None:
        super().__init__(parent)
        self.title("Bulk Copy")
        self.geometry("1220x560")
        self.minsize(1080, 480)
        self.transient(parent)
        self.grab_set()
        self._undo_manager = EntryUndoManager()

        self.mode = mode
        self._on_run = on_run
        if mode == "direct_upload":
            self.columns = self.DIRECT_UPLOAD_COLUMNS
            self.header_aliases = self.DIRECT_UPLOAD_HEADER_ALIASES
            self.description_text = (
                "Optional bulk mode: select source files, then export CSV template to fill Desired columns and import it back."
            )
            self.starter_row = None
            self.run_action_label = "Run Bulk Upload"
            self.template_button_label = "Convert to CSV Template..."
        else:
            self.columns = self.S3_COPY_COLUMNS
            self.header_aliases = self.S3_COPY_HEADER_ALIASES
            self.description_text = (
                "Optional bulk mode: one row per title. Fill columns, then run all rows in one copy job."
            )
            self.starter_row = ("Title 1", "", "", "", "", "")
            self.run_action_label = "Run Bulk Copy"
            self.template_button_label = "Download Template..."

        main = ttk.Frame(self, padding=12)
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(1, weight=1)

        ttk.Label(
            main,
            text=self.description_text,
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        table_frame = ttk.Frame(main)
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(
            table_frame,
            columns=[column_id for column_id, _, _ in self.columns],
            show="headings",
            selectmode="browse",
        )
        for column_id, heading, width in self.columns:
            self.tree.heading(column_id, text=heading)
            self.tree.column(column_id, width=width, minwidth=120, anchor="w")

        vertical_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        horizontal_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vertical_scroll.set, xscrollcommand=horizontal_scroll.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vertical_scroll.grid(row=0, column=1, sticky="ns")
        horizontal_scroll.grid(row=1, column=0, sticky="ew")

        controls = ttk.Frame(main)
        controls.grid(row=2, column=0, sticky="ew", pady=(10, 8))
        controls.columnconfigure(6, weight=1)

        ttk.Button(controls, text="Add Row", command=self._add_row).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(controls, text="Duplicate Row", command=self._duplicate_selected_row).grid(
            row=0, column=1, padx=(0, 6)
        )
        ttk.Button(controls, text="Delete Row", command=self._delete_selected_row).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(controls, text="Import Spreadsheet...", command=self._import_spreadsheet).grid(
            row=0, column=3, padx=(0, 6)
        )
        ttk.Button(controls, text=self.template_button_label, command=self._download_template).grid(
            row=0, column=4, padx=(0, 6)
        )
        if self.mode == "direct_upload":
            ttk.Button(controls, text="Select Source Files...", command=self._select_source_files).grid(
                row=0, column=5, padx=(0, 6)
            )

        editor = ttk.LabelFrame(main, text="Selected Row Editor")
        editor.grid(row=3, column=0, sticky="ew", pady=(0, 8))
        editor.columnconfigure(0, weight=1)
        editor.columnconfigure(2, weight=1)
        editor.columnconfigure(4, weight=1)

        self.editor_vars: dict[str, tk.StringVar] = {}
        for index, (column_id, heading, _) in enumerate(self.columns):
            self.editor_vars[column_id] = tk.StringVar()
            self._undo_manager.register_variable(self.editor_vars[column_id])
            editor_row = (index // 3) * 2
            editor_col = (index % 3) * 2
            ttk.Label(editor, text=heading).grid(row=editor_row, column=editor_col, sticky="w", padx=8, pady=(6, 2))
            entry = ttk.Entry(editor, textvariable=self.editor_vars[column_id])
            entry.grid(row=editor_row + 1, column=editor_col, sticky="ew", padx=8, pady=(0, 6))
            ttk.Button(
                editor,
                text="Clear",
                width=CLEAR_BUTTON_WIDTH,
                command=lambda value=self.editor_vars[column_id]: value.set(""),
            ).grid(row=editor_row + 1, column=editor_col + 1, sticky="w", padx=(0, 8), pady=(0, 6))

        button_row = ttk.Frame(main)
        button_row.grid(row=4, column=0, sticky="e")
        ttk.Button(button_row, text="Apply Row Changes", command=self._apply_editor_to_selected).grid(
            row=0, column=0, padx=(0, 8)
        )
        ttk.Button(button_row, text="Cancel", command=self.destroy).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(button_row, text=self.run_action_label, command=self._run_bulk_copy).grid(row=0, column=2)

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        if self.starter_row:
            self._add_row(self.starter_row)
        self._bind_standard_shortcuts()

    def _bind_standard_shortcuts(self) -> None:
        for entry_class in ("Entry", "TEntry"):
            self.bind_class(entry_class, "<Command-z>", self._handle_undo_shortcut, add="+")
            self.bind_class(entry_class, "<Command-Z>", self._handle_redo_shortcut, add="+")
            self.bind_class(entry_class, "<Control-z>", self._handle_undo_shortcut, add="+")
            self.bind_class(entry_class, "<Control-Z>", self._handle_redo_shortcut, add="+")
        self.bind_class("Text", "<Command-z>", self._handle_undo_shortcut, add="+")
        self.bind_class("Text", "<Command-Z>", self._handle_redo_shortcut, add="+")
        self.bind_class("Text", "<Control-z>", self._handle_undo_shortcut, add="+")
        self.bind_class("Text", "<Control-Z>", self._handle_redo_shortcut, add="+")

    def _handle_undo_shortcut(self, event) -> str | None:
        return self._undo_manager.undo_from_widget(event.widget)

    def _handle_redo_shortcut(self, event) -> str | None:
        return self._undo_manager.redo_from_widget(event.widget)

    def _make_default_title(self) -> str:
        if self.mode == "direct_upload":
            return str(len(self.tree.get_children()) + 1)
        return f"Title {len(self.tree.get_children()) + 1}"

    def _add_row(self, values: tuple[str, ...] | None = None) -> None:
        row_values = values or (self._make_default_title(),) + ("",) * (len(self.columns) - 1)
        item_id = self.tree.insert("", "end", values=row_values)
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self._populate_editor_from_item(item_id)

    @staticmethod
    def _normalize_header(header: str) -> str:
        return " ".join(header.strip().lower().replace("_", " ").split())

    @staticmethod
    def _to_string(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_csv_rows(self, file_path: str) -> list[dict[str, str]]:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as file_handle:
            sample = file_handle.read(4096)
            file_handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel

            reader = csv.DictReader(file_handle, dialect=dialect)
            if not reader.fieldnames:
                raise ValueError("No header row found in CSV file.")

            return self._map_rows(reader.fieldnames, reader)

    def _parse_xlsx_rows(self, file_path: str) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError("Excel import requires openpyxl. Please install app dependencies.") from error

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if not header_row:
                raise ValueError("No header row found in Excel file.")

            headers = [self._to_string(value) for value in header_row]
            raw_rows = []
            for row in iterator:
                raw_row = {}
                for index, cell_value in enumerate(row):
                    key = headers[index] if index < len(headers) else ""
                    if key:
                        raw_row[key] = self._to_string(cell_value)
                raw_rows.append(raw_row)

            return self._map_rows(headers, raw_rows)
        finally:
            workbook.close()

    def _map_rows(self, headers: list[str], raw_rows) -> list[dict[str, str]]:
        def resolve_column_id(header: str) -> str | None:
            normalized = self._normalize_header(header)
            valid_column_ids = {column_id for column_id, _, _ in self.columns}
            if normalized in self.header_aliases:
                mapped_id = self.header_aliases[normalized]
                return mapped_id if mapped_id in valid_column_ids else None

            if header in valid_column_ids:
                return header
            return None

        column_map: dict[str, str] = {}
        for header in headers:
            column_id = resolve_column_id(self._to_string(header))
            if column_id:
                column_map[self._to_string(header)] = column_id

        if not column_map:
            expected_columns = ", ".join(heading for _, heading, _ in self.columns)
            raise ValueError(
                f"No recognized columns found. Expected columns like: {expected_columns}"
            )

        mapped_rows: list[dict[str, str]] = []
        for index, raw_row in enumerate(raw_rows, start=1):
            mapped = {column_id: "" for column_id, _, _ in self.columns}
            for header, value in raw_row.items():
                normalized_header = self._to_string(header)
                column_id = column_map.get(normalized_header)
                if column_id:
                    mapped[column_id] = self._to_string(value)

            if not mapped["title"]:
                mapped["title"] = str(index) if self.mode == "direct_upload" else f"Title {index}"

            if any(value for key, value in mapped.items() if key != "title"):
                mapped_rows.append(mapped)

        return mapped_rows

    def _replace_rows(self, mapped_rows: list[dict[str, str]]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        for row in mapped_rows:
            values = tuple(row.get(column_id, "") for column_id, _, _ in self.columns)
            self._add_row(values)

    @staticmethod
    def _normalized_title(value: str) -> str:
        return " ".join(value.strip().lower().split())

    def _merge_rows(self, mapped_rows: list[dict[str, str]]) -> tuple[int, int]:
        existing_item_ids = list(self.tree.get_children())
        existing_by_title: dict[str, str] = {}
        for item_id in existing_item_ids:
            values = self.tree.item(item_id, "values")
            title = str(values[0]).strip() if values else ""
            normalized = self._normalized_title(title)
            if normalized and normalized not in existing_by_title:
                existing_by_title[normalized] = item_id

        updated_count = 0
        added_count = 0
        for row in mapped_rows:
            title = str(row.get("title", "")).strip()
            normalized = self._normalized_title(title)
            item_id = existing_by_title.get(normalized) if normalized else None

            if item_id:
                existing_values = self.tree.item(item_id, "values")
                merged_values = []
                for index, (column_id, _, _) in enumerate(self.columns):
                    existing_value = str(existing_values[index]) if index < len(existing_values) else ""
                    imported_value = str(row.get(column_id, "")).strip()
                    merged_values.append(imported_value if imported_value else existing_value)
                self.tree.item(item_id, values=tuple(merged_values))
                updated_count += 1
            else:
                values = tuple(row.get(column_id, "") for column_id, _, _ in self.columns)
                self._add_row(values)
                added_count += 1
        return updated_count, added_count

    def _template_file_name(self) -> str:
        if self.mode == "direct_upload":
            return "bulk_upload_title_mapping.csv"
        return "bulk_copy_template.csv"

    def _build_template_rows(self) -> list[tuple[str, ...]]:
        if self.mode == "direct_upload":
            return [
                ("1", "/Users/your.name/Downloads/source_video_1.mp4", "folder1/folder2", "new_video_name.mp4"),
                ("2", "/Users/your.name/Downloads/source_video_2.mp4", "folder1/folder2", "new_video_name_2.mp4"),
            ]
        return [
            ("Title 1", "current_video_name.mp4", "", "folder1/folder2", "new_video_name.mp4", ""),
            ("Title 2", "current_video_name_2.mp4", "current_video_name_2.vtt", "folder1/folder2", "new_video_name_2.mp4", "new_video_name_2.vtt"),
        ]

    def _download_template(self) -> None:
        save_path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Template",
            defaultextension=".csv",
            initialfile=self._template_file_name(),
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*")),
        )
        if not save_path:
            return

        headers = [heading for _, heading, _ in self.columns]
        rows_to_write = self._build_template_rows()
        if self.mode == "direct_upload":
            populated_rows: list[tuple[str, ...]] = []
            for row in self._collect_rows():
                has_payload = any(
                    str(row.get(column_id, "")).strip()
                    for column_id, _, _ in self.columns
                    if column_id != "title"
                )
                if has_payload:
                    populated_rows.append(
                        tuple(str(row.get(column_id, "")).strip() for column_id, _, _ in self.columns)
                    )

            if not populated_rows:
                messagebox.showerror(
                    "Convert to CSV Template",
                    "Select source files first, then click Convert to CSV Template.",
                    parent=self,
                )
                return
            rows_to_write = populated_rows

        try:
            with open(save_path, "w", encoding="utf-8", newline="") as file_handle:
                writer = csv.writer(file_handle)
                writer.writerow(headers)
                writer.writerows(rows_to_write)
            if self.mode == "direct_upload":
                messagebox.showinfo(
                    "CSV Template Saved",
                    (
                        f"CSV template saved to:\n{save_path}\n\n"
                        "Fill Desired Move Folder and Desired Name columns, save, then import spreadsheet."
                    ),
                    parent=self,
                )
            else:
                messagebox.showinfo("Template Saved", f"Template saved to:\n{save_path}", parent=self)
        except Exception as error:  # pylint: disable=broad-except
            messagebox.showerror("Template Save Failed", str(error), parent=self)

    def _import_spreadsheet(self) -> None:
        file_path = filedialog.askopenfilename(
            parent=self,
            title="Import Bulk Copy File",
            filetypes=(
                ("Spreadsheet files", "*.xlsx *.csv"),
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ),
        )
        if not file_path:
            return

        try:
            suffix = Path(file_path).suffix.lower()
            if suffix == ".xlsx":
                mapped_rows = self._parse_xlsx_rows(file_path)
            elif suffix == ".csv":
                mapped_rows = self._parse_csv_rows(file_path)
            else:
                raise ValueError("Unsupported file type. Please use .xlsx or .csv.")

            if not mapped_rows:
                raise ValueError("No populated rows found in the selected file.")

            existing_rows = list(self.tree.get_children())
            if not existing_rows:
                self._replace_rows(mapped_rows)
                messagebox.showinfo(
                    "Bulk Copy",
                    f"Imported {len(mapped_rows)} row(s). Review values, then click {self.run_action_label}.",
                    parent=self,
                )
                return

            if self.mode == "direct_upload":
                prompt = (
                    "Merge imported rows with existing rows by Title #?\n\n"
                    "Yes: Update matching titles (blank imported cells keep current values, including Local File Path).\n"
                    "No: Replace all existing rows."
                )
            else:
                prompt = (
                    "Merge imported rows with existing rows by Title?\n\n"
                    "Yes: Update matching titles (blank imported cells keep current values).\n"
                    "No: Replace all existing rows."
                )

            merge_choice = messagebox.askyesnocancel(
                "Import Mode",
                prompt,
                parent=self,
            )
            if merge_choice is None:
                return
            if merge_choice:
                updated_count, added_count = self._merge_rows(mapped_rows)
                messagebox.showinfo(
                    "Bulk Copy",
                    (
                        f"Import complete.\n\nUpdated rows: {updated_count}\nAdded rows: {added_count}\n\n"
                        f"Review values, then click {self.run_action_label}."
                    ),
                    parent=self,
                )
            else:
                self._replace_rows(mapped_rows)
                messagebox.showinfo(
                    "Bulk Copy",
                    f"Imported {len(mapped_rows)} row(s). Review values, then click {self.run_action_label}.",
                    parent=self,
                )
        except Exception as error:  # pylint: disable=broad-except
            messagebox.showerror("Bulk Copy Import Failed", str(error), parent=self)

    def _duplicate_selected_row(self) -> None:
        selected_item = self._get_selected_item()
        if not selected_item:
            return
        values = tuple(self.tree.item(selected_item, "values"))
        title = str(values[0]).strip() if values else ""
        if self.mode == "direct_upload":
            duplicate_title = self._make_default_title()
        else:
            duplicate_title = f"{title} Copy" if title else self._make_default_title()
        duplicated = (duplicate_title,) + tuple(values[1:])
        self._add_row(duplicated)

    def _delete_selected_row(self) -> None:
        selected_item = self._get_selected_item()
        if not selected_item:
            return
        self.tree.delete(selected_item)
        remaining = self.tree.get_children()
        if remaining:
            self.tree.selection_set(remaining[0])
            self.tree.focus(remaining[0])
            self._populate_editor_from_item(remaining[0])
        else:
            for var in self.editor_vars.values():
                var.set("")

    def _get_selected_item(self) -> str | None:
        selected = self.tree.selection()
        if not selected:
            return None
        return selected[0]

    def _on_tree_select(self, _event) -> None:
        selected_item = self._get_selected_item()
        if not selected_item:
            return
        self._populate_editor_from_item(selected_item)

    def _populate_editor_from_item(self, item_id: str) -> None:
        values = self.tree.item(item_id, "values")
        for index, (column_id, _, _) in enumerate(self.columns):
            self.editor_vars[column_id].set(str(values[index]) if index < len(values) else "")

    def _apply_editor_to_selected(self) -> None:
        selected_item = self._get_selected_item()
        if not selected_item:
            messagebox.showerror("Bulk Copy", "Select a row first.", parent=self)
            return

        values = tuple(self.editor_vars[column_id].get() for column_id, _, _ in self.columns)
        self.tree.item(selected_item, values=values)

    def _collect_rows(self) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            row: dict[str, str] = {}
            for index, (column_id, _, _) in enumerate(self.columns):
                row[column_id] = str(values[index]) if index < len(values) else ""
            rows.append(row)
        return rows

    def _run_bulk_copy(self) -> None:
        # Do not auto-apply editor values on run. This avoids accidental row mutation
        # when the editor has stale or partial values.
        rows = self._collect_rows()
        if not rows:
            messagebox.showerror("Bulk Copy", "Add at least one row before running bulk copy.", parent=self)
            return

        started = self._on_run(rows)
        if started:
            self.destroy()

    def _select_source_files(self) -> None:
        selected_files = filedialog.askopenfilenames(
            parent=self,
            title="Select Source Files for Bulk Upload",
        )
        if not selected_files:
            return

        selected_item = self._get_selected_item()
        default_move_folder = ""
        column_order = [column_id for column_id, _, _ in self.columns]
        move_folder_index = column_order.index("desired_move_folder") if "desired_move_folder" in column_order else -1
        if selected_item:
            selected_values = self.tree.item(selected_item, "values")
            if move_folder_index >= 0 and len(selected_values) > move_folder_index:
                default_move_folder = str(selected_values[move_folder_index]).strip()

        next_title_number = self._next_direct_upload_title_number()
        for offset, file_path in enumerate(selected_files):
            file_name = Path(file_path).name
            title = str(next_title_number + offset)
            row_values_by_column = {
                "title": title,
                "local_file_path": str(file_path),
                "desired_move_folder": default_move_folder,
                "desired_name": file_name,
            }
            row_values = tuple(row_values_by_column.get(column_id, "") for column_id in column_order)
            self._add_row(row_values)

    def _next_direct_upload_title_number(self) -> int:
        highest_number = 0
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            if not values:
                continue
            title_value = str(values[0]).strip()
            digits = "".join(character for character in title_value if character.isdigit())
            if digits:
                highest_number = max(highest_number, int(digits))
        return highest_number + 1


class S3CopyApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.minsize(740, 700)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        self.config = load_config()
        self.use_session_only_credentials = False
        self.session_credentials: AwsCredentials | None = None
        self.keychain_credentials: AwsCredentials | None = None
        self._keychain_credentials_loaded = False
        self._ui_queue: queue.Queue = queue.Queue()
        self._running = False
        self._closing = False
        self._pause_requested = False
        self._pause_active = False
        self._undo_manager = EntryUndoManager()
        self.simplified_bulk_dry_run_button: ttk.Button | None = None

        self.current_file_name_var = tk.StringVar()
        self.current_caption_name_var = tk.StringVar()
        self.local_file_path_var = tk.StringVar()
        self.local_caption_path_var = tk.StringVar()
        self.rename_current_path_var = tk.StringVar()
        self.rename_current_name_var = tk.StringVar()
        self.rename_desired_name_var = tk.StringVar()
        self.simplified_bulk_csv_path_var = tk.StringVar()
        self.simplified_bulk_summary_var = tk.StringVar(
            value="Load a CSV with source_uri and destination_uri columns. Resume details appear here if a prior run exists."
        )
        self.inventory_path_var = tk.StringVar()
        self.inventory_summary_var = tk.StringVar(
            value="Enter an S3 bucket or prefix URI to export a CSV inventory of everything under that location."
        )
        self.folder_copy_source_uri_var = tk.StringVar()
        self.folder_copy_dest_uri_var = tk.StringVar()
        self.folder_copy_summary_var = tk.StringVar(
            value="Provide source and destination S3 folder URIs. Relative paths under the source folder will be preserved."
        )
        self.bulk_folder_csv_path_var = tk.StringVar()
        self.bulk_folder_summary_var = tk.StringVar(
            value="Load a CSV with source_folder_uri and destination_folder_uri columns to copy multiple folder pairs."
        )
        self.desired_move_folder_var = tk.StringVar()
        self.desired_name_var = tk.StringVar()
        self.desired_caption_name_var = tk.StringVar()

        self.source_preview_var = tk.StringVar(value="")
        self.dest_preview_var = tk.StringVar(value="")
        self.source_caption_preview_var = tk.StringVar(value="")
        self.dest_caption_preview_var = tk.StringVar(value="")

        self._build_menu()
        self._build_layout()
        self._bind_field_events()
        self._bind_standard_shortcuts()

        # Tkinter objects can crash if their GC finalizers run on non-main threads.
        # Keep cyclic GC disabled globally and collect only on the UI thread.
        gc.disable()
        self.root.after(100, self._process_ui_queue)
        self.root.after(10_000, self._collect_gc_on_ui_thread)
        self._refresh_preview()
        self._log_credential_mode_on_startup()

    def _build_menu(self) -> None:
        menu_bar = tk.Menu(self.root)
        self.settings_menu = tk.Menu(menu_bar, tearoff=0)
        self.settings_menu.add_command(label="New Instance", command=self.open_new_instance)
        self.settings_menu.add_command(label="Bulk Copy...", command=self.open_bulk_copy)
        self.bulk_menu_index = 1
        self.settings_menu.add_command(label="Download Bulk Copy Template...", command=self.download_bulk_copy_template)
        self.settings_menu.add_command(
            label="Download Bulk Upload CSV Template...",
            command=self.download_bulk_upload_template,
        )
        self.settings_menu.add_command(label="Settings...", command=self.open_settings)
        self.settings_menu.add_separator()
        self.settings_menu.add_command(label="Quit", command=self.root.destroy)
        menu_bar.add_cascade(label="App", menu=self.settings_menu)
        self.root.config(menu=menu_bar)

    def open_new_instance(self) -> None:
        try:
            if getattr(sys, "frozen", False):
                launch_command = [sys.executable]
                launch_cwd = None
            else:
                launch_command = [sys.executable, "-m", "s3_copy_desktop_app.app"]
                launch_cwd = str(Path(__file__).resolve().parents[1])

            subprocess.Popen(launch_command, cwd=launch_cwd)
            self._append_log("Launched a new app instance.")
        except Exception as error:  # pylint: disable=broad-except
            self._append_log(f"Could not launch a new instance: {error}")
            messagebox.showerror("New Instance", f"Could not launch a new instance: {error}", parent=self.root)

    def _make_clear_button(
        self,
        parent: tk.Widget,
        variable: tk.StringVar,
        bg: str,
        row: int,
        column: int,
        pady,
    ) -> tk.Button:
        button = tk.Button(
            parent,
            text="Clear",
            command=lambda value=variable: value.set(""),
            bg=bg,
            fg=SECTION_TEXT_COLOR,
            activebackground=bg,
            activeforeground=SECTION_TEXT_COLOR,
            relief="flat",
            borderwidth=1,
            highlightthickness=0,
            padx=10,
            pady=2,
            width=5,
            takefocus=False,
        )
        button.grid(row=row, column=column, sticky="e", padx=(8, 0), pady=pady)
        return button

    def _make_entry(
        self,
        parent: tk.Widget,
        variable: tk.StringVar,
        bg: str,
        fg: str,
        row: int,
        column: int = 1,
        pady=(0, 6),
        state: str = "normal",
        readonlybackground: str | None = None,
    ) -> tk.Entry:
        entry_kwargs = {
            "textvariable": variable,
            "bg": bg,
            "fg": fg,
            "relief": "solid",
            "borderwidth": 1,
            "state": state,
        }
        if readonlybackground is not None:
            entry_kwargs["readonlybackground"] = readonlybackground
        entry = tk.Entry(parent, **entry_kwargs)
        if state == "normal":
            self._undo_manager.register_variable(variable)
        entry.grid(row=row, column=column, sticky="ew", pady=pady)
        return entry

    def _bind_standard_shortcuts(self) -> None:
        for entry_class in ("Entry", "TEntry"):
            self.root.bind_class(entry_class, "<Command-z>", self._handle_undo_shortcut, add="+")
            self.root.bind_class(entry_class, "<Command-Z>", self._handle_redo_shortcut, add="+")
            self.root.bind_class(entry_class, "<Control-z>", self._handle_undo_shortcut, add="+")
            self.root.bind_class(entry_class, "<Control-Z>", self._handle_redo_shortcut, add="+")
        self.root.bind_class("Text", "<Command-z>", self._handle_undo_shortcut, add="+")
        self.root.bind_class("Text", "<Command-Z>", self._handle_redo_shortcut, add="+")
        self.root.bind_class("Text", "<Control-z>", self._handle_undo_shortcut, add="+")
        self.root.bind_class("Text", "<Control-Z>", self._handle_redo_shortcut, add="+")

    def _handle_undo_shortcut(self, event) -> str | None:
        return self._undo_manager.undo_from_widget(event.widget)

    def _handle_redo_shortcut(self, event) -> str | None:
        return self._undo_manager.redo_from_widget(event.widget)

    def _build_layout(self) -> None:
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(2, weight=1)

        input_frame = ttk.LabelFrame(main, text="Input", padding=10)
        input_frame.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        input_frame.columnconfigure(0, weight=1)

        top_row = ttk.Frame(input_frame)
        top_row.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        top_row.columnconfigure(0, weight=1)
        ttk.Label(top_row, text="Need to update bucket paths or AWS credentials?").grid(row=0, column=0, sticky="w")
        self.bulk_copy_button = ttk.Button(top_row, text="Bulk Copy...", command=self.open_bulk_copy)
        self.bulk_copy_button.grid(row=0, column=1, sticky="e", padx=(0, 8))
        ttk.Button(top_row, text="Settings...", command=self.open_settings).grid(row=0, column=2, sticky="e")

        self.mode_notebook = ttk.Notebook(input_frame)
        self.mode_notebook.grid(row=1, column=0, sticky="ew", pady=(0, 4))
        self.mode_notebook.configure(height=188)
        self.s3_mode_frame = ttk.Frame(self.mode_notebook, padding=6)
        self.direct_mode_frame = ttk.Frame(self.mode_notebook, padding=6)
        self.rename_mode_frame = ttk.Frame(self.mode_notebook, padding=6)
        self.simplified_bulk_mode_frame = ttk.Frame(self.mode_notebook, padding=6)
        self.inventory_mode_frame = ttk.Frame(self.mode_notebook, padding=6)
        self.folder_copy_mode_frame = ttk.Frame(self.mode_notebook, padding=6) if POWER_MODE else None
        self.bulk_folder_copy_mode_frame = ttk.Frame(self.mode_notebook, padding=6) if POWER_MODE else None
        self.mode_notebook.add(self.s3_mode_frame, text="S3 Copy")
        self.mode_notebook.add(self.direct_mode_frame, text="Direct Upload")
        self.mode_notebook.add(self.rename_mode_frame, text="Rename in Destination")
        self.mode_notebook.add(self.simplified_bulk_mode_frame, text="Simplified Bulk Copy")
        self.mode_notebook.add(self.inventory_mode_frame, text="Inventory")
        if self.folder_copy_mode_frame is not None:
            self.mode_notebook.add(self.folder_copy_mode_frame, text="Folder Copy")
        if self.bulk_folder_copy_mode_frame is not None:
            self.mode_notebook.add(self.bulk_folder_copy_mode_frame, text="Bulk Folder Copy")
        self.s3_mode_frame.columnconfigure(0, weight=1)
        self.direct_mode_frame.columnconfigure(0, weight=1)
        self.rename_mode_frame.columnconfigure(0, weight=1)
        self.simplified_bulk_mode_frame.columnconfigure(0, weight=1)
        self.inventory_mode_frame.columnconfigure(0, weight=1)
        if self.folder_copy_mode_frame is not None:
            self.folder_copy_mode_frame.columnconfigure(0, weight=1)
        if self.bulk_folder_copy_mode_frame is not None:
            self.bulk_folder_copy_mode_frame.columnconfigure(0, weight=1)

        s3_current_block = tk.Frame(
            self.s3_mode_frame,
            bg=CURRENT_BLOCK_BG,
            bd=1,
            relief="groove",
            padx=10,
            pady=6,
        )
        s3_current_block.grid(row=0, column=0, sticky="ew")
        s3_current_block.columnconfigure(1, weight=1)
        tk.Label(
            s3_current_block,
            text="Current File Name",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            s3_current_block,
            self.current_file_name_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=0,
        )
        self._make_clear_button(s3_current_block, self.current_file_name_var, CURRENT_FIELD_BG, row=0, column=2, pady=(0, 6))
        tk.Label(
            s3_current_block,
            text="Current Caption Name (optional)",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(
            row=1, column=0, sticky="w", pady=(0, 6), padx=(0, 10)
        )
        self._make_entry(
            s3_current_block,
            self.current_caption_name_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=1,
        )
        self._make_clear_button(
            s3_current_block,
            self.current_caption_name_var,
            CURRENT_FIELD_BG,
            row=1,
            column=2,
            pady=(0, 6),
        )

        direct_current_block = tk.Frame(
            self.direct_mode_frame,
            bg=CURRENT_BLOCK_BG,
            bd=1,
            relief="groove",
            padx=10,
            pady=6,
        )
        direct_current_block.grid(row=0, column=0, sticky="ew")
        direct_current_block.columnconfigure(1, weight=1)
        tk.Label(
            direct_current_block,
            text="Local File (source)",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(
            row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10)
        )
        self._make_entry(
            direct_current_block,
            self.local_file_path_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=0,
        )
        self._make_clear_button(
            direct_current_block,
            self.local_file_path_var,
            CURRENT_FIELD_BG,
            row=0,
            column=2,
            pady=(0, 6),
        )
        tk.Button(
            direct_current_block,
            text="Browse...",
            command=self._browse_local_file,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            activebackground=CURRENT_FIELD_BG,
            activeforeground=SECTION_TEXT_COLOR,
            relief="flat",
            borderwidth=1,
            highlightthickness=0,
            padx=10,
            pady=2,
            takefocus=False,
        ).grid(
            row=0, column=3, sticky="e", padx=(8, 0), pady=(0, 6)
        )
        tk.Label(
            direct_current_block,
            text="Local Caption (optional)",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(
            row=1, column=0, sticky="w", pady=(0, 6), padx=(0, 10)
        )
        self._make_entry(
            direct_current_block,
            self.local_caption_path_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=1,
        )
        self._make_clear_button(
            direct_current_block,
            self.local_caption_path_var,
            CURRENT_FIELD_BG,
            row=1,
            column=2,
            pady=(0, 6),
        )
        tk.Button(
            direct_current_block,
            text="Browse...",
            command=self._browse_local_caption,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            activebackground=CURRENT_FIELD_BG,
            activeforeground=SECTION_TEXT_COLOR,
            relief="flat",
            borderwidth=1,
            highlightthickness=0,
            padx=10,
            pady=2,
            takefocus=False,
        ).grid(
            row=1, column=3, sticky="e", padx=(8, 0), pady=(0, 6)
        )

        rename_block = tk.Frame(
            self.rename_mode_frame,
            bg=CURRENT_BLOCK_BG,
            bd=1,
            relief="groove",
            padx=10,
            pady=6,
        )
        rename_block.grid(row=0, column=0, sticky="ew")
        rename_block.columnconfigure(1, weight=1)
        tk.Label(
            rename_block,
            text="Current File Path (after default destination prefix)",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            rename_block,
            self.rename_current_path_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=0,
        )
        self._make_clear_button(rename_block, self.rename_current_path_var, CURRENT_FIELD_BG, row=0, column=2, pady=(0, 6))
        tk.Label(
            rename_block,
            text="Current File Name",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=1, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            rename_block,
            self.rename_current_name_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=1,
        )
        self._make_clear_button(rename_block, self.rename_current_name_var, CURRENT_FIELD_BG, row=1, column=2, pady=(0, 6))
        tk.Label(
            rename_block,
            text="Desired File Name",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=2, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            rename_block,
            self.rename_desired_name_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=2,
        )
        self._make_clear_button(rename_block, self.rename_desired_name_var, CURRENT_FIELD_BG, row=2, column=2, pady=(0, 6))
        tk.Label(
            rename_block,
            text="Note: Renaming large .mov files may take up to 5 minutes. Smaller files should be faster.",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            wraplength=640,
            justify="left",
        ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(2, 4))

        simplified_bulk_block = tk.Frame(
            self.simplified_bulk_mode_frame,
            bg=CURRENT_BLOCK_BG,
            bd=1,
            relief="groove",
            padx=10,
            pady=8,
        )
        simplified_bulk_block.grid(row=0, column=0, sticky="ew")
        simplified_bulk_block.columnconfigure(1, weight=1)
        tk.Label(
            simplified_bulk_block,
            text="CSV File",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self.simplified_bulk_csv_entry = self._make_entry(
            simplified_bulk_block,
            self.simplified_bulk_csv_path_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=0,
        )
        self._make_clear_button(
            simplified_bulk_block,
            self.simplified_bulk_csv_path_var,
            CURRENT_FIELD_BG,
            row=0,
            column=2,
            pady=(0, 6),
        )
        tk.Button(
            simplified_bulk_block,
            text="Browse...",
            command=self._browse_simplified_bulk_csv,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            activebackground=CURRENT_FIELD_BG,
            activeforeground=SECTION_TEXT_COLOR,
            relief="flat",
            borderwidth=1,
            highlightthickness=0,
            padx=8,
            pady=2,
            takefocus=False,
        ).grid(row=0, column=3, sticky="e", padx=(8, 0), pady=(0, 6))
        tk.Label(
            simplified_bulk_block,
            text="Summary",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=1, column=0, sticky="nw", padx=(0, 10))
        tk.Label(
            simplified_bulk_block,
            textvariable=self.simplified_bulk_summary_var,
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            justify="left",
            wraplength=640,
        ).grid(row=1, column=1, columnspan=3, sticky="w")
        tk.Label(
            simplified_bulk_block,
            text=(
                "CSV mode uses the full source_uri and destination_uri values from the file, ignores "
                "Source/Destination bucket and prefix settings for those rows, and "
                + (
                    "always runs a required dry run before any copy starts."
                    if SIMPLIFIED_BULK_REQUIRE_DRY_RUN
                    else "lets you run dry run optionally before copy starts."
                )
            ),
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            justify="left",
            wraplength=640,
        ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(8, 0))
        if not SIMPLIFIED_BULK_REQUIRE_DRY_RUN:
            self.simplified_bulk_dry_run_button = ttk.Button(
                simplified_bulk_block,
                text="Run Dry Run",
                command=self._on_simplified_bulk_dry_run_clicked,
            )
            self.simplified_bulk_dry_run_button.grid(row=3, column=3, sticky="e", pady=(8, 0))

        inventory_block = tk.Frame(
            self.inventory_mode_frame,
            bg=CURRENT_BLOCK_BG,
            bd=1,
            relief="groove",
            padx=10,
            pady=8,
        )
        inventory_block.grid(row=0, column=0, sticky="ew")
        inventory_block.columnconfigure(1, weight=1)
        tk.Label(
            inventory_block,
            text="S3 Bucket or Prefix URI",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            inventory_block,
            self.inventory_path_var,
            bg=CURRENT_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=0,
        )
        self._make_clear_button(
            inventory_block,
            self.inventory_path_var,
            CURRENT_FIELD_BG,
            row=0,
            column=2,
            pady=(0, 6),
        )
        tk.Label(
            inventory_block,
            text="Summary",
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=1, column=0, sticky="nw", padx=(0, 10))
        tk.Label(
            inventory_block,
            textvariable=self.inventory_summary_var,
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            justify="left",
            wraplength=640,
        ).grid(row=1, column=1, columnspan=2, sticky="w")
        tk.Label(
            inventory_block,
            text=(
                "Exports a CSV listing every object under the bucket/prefix you enter. "
                "This is read-only and does not copy, rename, or delete anything."
            ),
            bg=CURRENT_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            justify="left",
            wraplength=640,
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(8, 0))

        if self.folder_copy_mode_frame is not None:
            folder_copy_block = tk.Frame(
                self.folder_copy_mode_frame,
                bg=CURRENT_BLOCK_BG,
                bd=1,
                relief="groove",
                padx=10,
                pady=8,
            )
            folder_copy_block.grid(row=0, column=0, sticky="ew")
            folder_copy_block.columnconfigure(1, weight=1)
            tk.Label(
                folder_copy_block,
                text="Source Folder S3 URI",
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
            ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
            self._make_entry(
                folder_copy_block,
                self.folder_copy_source_uri_var,
                bg=CURRENT_FIELD_BG,
                fg=SECTION_TEXT_COLOR,
                row=0,
            )
            self._make_clear_button(
                folder_copy_block,
                self.folder_copy_source_uri_var,
                CURRENT_FIELD_BG,
                row=0,
                column=2,
                pady=(0, 6),
            )
            tk.Label(
                folder_copy_block,
                text="Destination Folder S3 URI",
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
            ).grid(row=1, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
            self._make_entry(
                folder_copy_block,
                self.folder_copy_dest_uri_var,
                bg=CURRENT_FIELD_BG,
                fg=SECTION_TEXT_COLOR,
                row=1,
            )
            self._make_clear_button(
                folder_copy_block,
                self.folder_copy_dest_uri_var,
                CURRENT_FIELD_BG,
                row=1,
                column=2,
                pady=(0, 6),
            )
            tk.Label(
                folder_copy_block,
                text="Summary",
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
            ).grid(row=2, column=0, sticky="nw", padx=(0, 10))
            tk.Label(
                folder_copy_block,
                textvariable=self.folder_copy_summary_var,
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
                justify="left",
                wraplength=640,
            ).grid(row=2, column=1, columnspan=2, sticky="w")
            tk.Label(
                folder_copy_block,
                text=(
                    "Copies every object under the source folder to the destination folder and preserves "
                    "the relative path for each object. Existing destination collisions are reviewed once at the end."
                ),
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
                justify="left",
                wraplength=640,
            ).grid(row=3, column=0, columnspan=3, sticky="w", pady=(8, 0))

        if self.bulk_folder_copy_mode_frame is not None:
            bulk_folder_block = tk.Frame(
                self.bulk_folder_copy_mode_frame,
                bg=CURRENT_BLOCK_BG,
                bd=1,
                relief="groove",
                padx=10,
                pady=8,
            )
            bulk_folder_block.grid(row=0, column=0, sticky="ew")
            bulk_folder_block.columnconfigure(1, weight=1)
            tk.Label(
                bulk_folder_block,
                text="CSV File",
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
            ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
            self.bulk_folder_csv_entry = self._make_entry(
                bulk_folder_block,
                self.bulk_folder_csv_path_var,
                bg=CURRENT_FIELD_BG,
                fg=SECTION_TEXT_COLOR,
                row=0,
            )
            self._make_clear_button(
                bulk_folder_block,
                self.bulk_folder_csv_path_var,
                CURRENT_FIELD_BG,
                row=0,
                column=2,
                pady=(0, 6),
            )
            tk.Button(
                bulk_folder_block,
                text="Browse...",
                command=self._browse_bulk_folder_csv,
                bg=CURRENT_FIELD_BG,
                fg=SECTION_TEXT_COLOR,
                activebackground=CURRENT_FIELD_BG,
                activeforeground=SECTION_TEXT_COLOR,
                relief="flat",
                borderwidth=1,
                highlightthickness=0,
                padx=8,
                pady=2,
                takefocus=False,
            ).grid(row=0, column=3, sticky="e", padx=(8, 0), pady=(0, 6))
            tk.Label(
                bulk_folder_block,
                text="Summary",
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
            ).grid(row=1, column=0, sticky="nw", padx=(0, 10))
            tk.Label(
                bulk_folder_block,
                textvariable=self.bulk_folder_summary_var,
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
                justify="left",
                wraplength=820,
            ).grid(row=1, column=1, columnspan=3, sticky="w")
            tk.Label(
                bulk_folder_block,
                text=(
                    "Load a CSV with source_folder_uri and destination_folder_uri columns. "
                    "Each row copies one source folder to one destination folder, preserving relative paths. "
                    "Destination overwrite candidates are still reviewed once at the end."
                ),
                bg=CURRENT_BLOCK_BG,
                fg=SECTION_TEXT_COLOR,
                justify="left",
                wraplength=820,
            ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(8, 0))

        self.desired_block = tk.Frame(
            input_frame,
            bg=DESIRED_BLOCK_BG,
            bd=1,
            relief="groove",
            padx=10,
            pady=6,
        )
        self.desired_block.grid(row=2, column=0, sticky="ew", pady=(0, 6))
        self.desired_block.columnconfigure(1, weight=1)
        tk.Label(
            self.desired_block,
            text="Desired Move Folder",
            bg=DESIRED_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=0, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            self.desired_block,
            self.desired_move_folder_var,
            bg=DESIRED_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=0,
        )
        self._make_clear_button(
            self.desired_block,
            self.desired_move_folder_var,
            DESIRED_FIELD_BG,
            row=0,
            column=2,
            pady=(0, 6),
        )
        tk.Label(
            self.desired_block,
            text="Desired Name",
            bg=DESIRED_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=1, column=0, sticky="w", pady=(0, 6), padx=(0, 10))
        self._make_entry(
            self.desired_block,
            self.desired_name_var,
            bg=DESIRED_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=1,
        )
        self._make_clear_button(
            self.desired_block,
            self.desired_name_var,
            DESIRED_FIELD_BG,
            row=1,
            column=2,
            pady=(0, 6),
        )
        self.desired_caption_label = tk.Label(
            self.desired_block,
            text="Desired Caption Name (optional)",
            bg=DESIRED_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        )
        self.desired_caption_label.grid(
            row=2, column=0, sticky="w", pady=(0, 8), padx=(0, 10)
        )
        self.desired_caption_entry = self._make_entry(
            self.desired_block,
            self.desired_caption_name_var,
            bg=DESIRED_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=2,
            pady=(0, 8),
        )
        self._make_clear_button(
            self.desired_block,
            self.desired_caption_name_var,
            DESIRED_FIELD_BG,
            row=2,
            column=2,
            pady=(0, 8),
        )

        action_row = ttk.Frame(input_frame)
        action_row.grid(row=3, column=0, sticky="e", pady=(0, 2))
        self.pause_button = ttk.Button(action_row, text="Pause", command=self._on_pause_resume_clicked, state="disabled")
        self.pause_button.grid(row=0, column=0, sticky="e", padx=(0, 8))
        self.copy_button = ttk.Button(action_row, text="Copy", command=self.on_copy_clicked)
        self.copy_button.grid(row=0, column=1, sticky="e")

        preview_frame = tk.LabelFrame(
            main,
            text="Resolved S3 Paths",
            bg=RESULTS_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            bd=1,
            relief="groove",
            padx=10,
            pady=8,
        )
        preview_frame.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        preview_frame.columnconfigure(1, weight=1)

        tk.Label(
            preview_frame,
            text="Read Only - for your information - do not input anything here.",
            bg=RESULTS_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(6, 4))

        tk.Label(preview_frame, text="Source", bg=RESULTS_BLOCK_BG, fg=SECTION_TEXT_COLOR).grid(
            row=1, column=0, sticky="w", padx=(10, 8), pady=(0, 4)
        )
        self._make_entry(
            preview_frame,
            self.source_preview_var,
            bg=RESULTS_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=1,
            state="readonly",
            readonlybackground=RESULTS_FIELD_BG,
        )

        tk.Label(preview_frame, text="Destination", bg=RESULTS_BLOCK_BG, fg=SECTION_TEXT_COLOR).grid(
            row=2, column=0, sticky="w", padx=(10, 8), pady=(0, 8)
        )
        self._make_entry(
            preview_frame,
            self.dest_preview_var,
            bg=RESULTS_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=2,
            pady=(0, 8),
            state="readonly",
            readonlybackground=RESULTS_FIELD_BG,
        )
        self.preview_source_caption_label = tk.Label(
            preview_frame,
            text="Source Caption",
            bg=RESULTS_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        )
        self.preview_source_caption_label.grid(
            row=3, column=0, sticky="w", padx=(10, 8), pady=(0, 4)
        )
        self.preview_source_caption_entry = self._make_entry(
            preview_frame,
            self.source_caption_preview_var,
            bg=RESULTS_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=3,
            pady=(0, 4),
            state="readonly",
            readonlybackground=RESULTS_FIELD_BG,
        )
        self.preview_dest_caption_label = tk.Label(
            preview_frame,
            text="Destination Caption",
            bg=RESULTS_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        )
        self.preview_dest_caption_label.grid(
            row=4, column=0, sticky="w", padx=(10, 8), pady=(0, 8)
        )
        self.preview_dest_caption_entry = self._make_entry(
            preview_frame,
            self.dest_caption_preview_var,
            bg=RESULTS_FIELD_BG,
            fg=SECTION_TEXT_COLOR,
            row=4,
            pady=(0, 8),
            state="readonly",
            readonlybackground=RESULTS_FIELD_BG,
        )
        tk.Label(
            preview_frame,
            text="Tip: highlight any resolved path and press Cmd+C to copy.",
            bg=RESULTS_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
        ).grid(row=5, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 6))

        log_frame = tk.LabelFrame(
            main,
            text="Status / Output",
            bg=RESULTS_BLOCK_BG,
            fg=SECTION_TEXT_COLOR,
            bd=1,
            relief="groove",
            padx=10,
            pady=8,
        )
        log_frame.grid(row=2, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_area = ScrolledText(
            log_frame,
            height=10,
            state="disabled",
            wrap="word",
            bg=RESULTS_FIELD_BG,
            fg="#0d2d4d",
            insertbackground="#0d2d4d",
            relief="flat",
        )
        self.log_area.grid(row=0, column=0, sticky="nsew")

        self._append_log(
            (
                "App started. Use S3 Copy, Direct Upload, Rename, Inventory"
                + (", Folder Copy, Bulk Folder Copy" if POWER_MODE else "")
                + ", or Simplified Bulk Copy, then click the main action button."
            )
        )

    def _bind_field_events(self) -> None:
        for variable in (
            self.current_file_name_var,
            self.current_caption_name_var,
            self.local_file_path_var,
            self.local_caption_path_var,
            self.rename_current_path_var,
            self.rename_current_name_var,
            self.rename_desired_name_var,
            self.simplified_bulk_csv_path_var,
            self.inventory_path_var,
            self.folder_copy_source_uri_var,
            self.folder_copy_dest_uri_var,
            self.bulk_folder_csv_path_var,
            self.desired_move_folder_var,
            self.desired_name_var,
            self.desired_caption_name_var,
        ):
            variable.trace_add("write", lambda *_: self._refresh_preview())
        self.mode_notebook.bind("<<NotebookTabChanged>>", lambda *_: self._refresh_preview())

    def _log_credential_mode_on_startup(self) -> None:
        if self.use_session_only_credentials:
            self._append_log("Credential mode is session-only. Saved credentials are not used unless you switch modes.")
            return
        self._append_log(
            f"Keychain/default AWS credential mode is active. Stored credentials will be loaded only when needed."
        )

    def open_settings(self) -> None:
        SettingsDialog(
            self.root,
            self.config,
            self.session_credentials if self.use_session_only_credentials else self.keychain_credentials,
            self.use_session_only_credentials,
            self._on_settings_saved,
        )

    def open_bulk_copy(self) -> None:
        if self._running:
            messagebox.showerror("Bulk Copy", "A copy is already running. Wait for it to finish first.", parent=self.root)
            return
        if self._is_rename_mode() or self._is_simplified_bulk_mode():
            messagebox.showerror(
                "Bulk Mode",
                "Bulk dialog mode is available for S3 Copy and Direct Upload tabs only.",
                parent=self.root,
            )
            return
        if self._is_direct_upload_mode():
            BulkCopyDialog(self.root, "direct_upload", self._start_bulk_direct_upload)
        else:
            BulkCopyDialog(self.root, "s3_copy", self._start_bulk_copy)

    def _browse_local_file(self) -> None:
        file_path = filedialog.askopenfilename(parent=self.root, title="Select Local Source File")
        if file_path:
            self.local_file_path_var.set(file_path)

    def _browse_local_caption(self) -> None:
        file_path = filedialog.askopenfilename(parent=self.root, title="Select Local Caption File")
        if file_path:
            self.local_caption_path_var.set(file_path)

    def _browse_simplified_bulk_csv(self) -> None:
        file_path = filedialog.askopenfilename(
            parent=self.root,
            title="Select Simplified Bulk Copy CSV",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*")),
        )
        if file_path:
            self.simplified_bulk_csv_path_var.set(file_path)
            self.root.after(10, self._show_simplified_bulk_csv_path_end)

    def _show_simplified_bulk_csv_path_end(self) -> None:
        try:
            self.simplified_bulk_csv_entry.icursor("end")
            self.simplified_bulk_csv_entry.xview_moveto(1.0)
        except Exception:  # pylint: disable=broad-except
            pass

    def _browse_bulk_folder_csv(self) -> None:
        file_path = filedialog.askopenfilename(
            parent=self.root,
            title="Select Bulk Folder Copy CSV",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*")),
        )
        if file_path:
            self.bulk_folder_csv_path_var.set(file_path)
            self.root.after(10, self._show_bulk_folder_csv_path_end)

    def _show_bulk_folder_csv_path_end(self) -> None:
        try:
            self.bulk_folder_csv_entry.icursor("end")
            self.bulk_folder_csv_entry.xview_moveto(1.0)
        except Exception:  # pylint: disable=broad-except
            pass

    @staticmethod
    def _simplified_bulk_checkpoint_path(csv_path: str) -> Path:
        normalized_path = str(Path(csv_path).expanduser().resolve())
        session_id = hashlib.sha256(normalized_path.encode("utf-8")).hexdigest()[:16]
        return SIMPLIFIED_BULK_CHECKPOINT_DIR / f"{session_id}.json"

    @staticmethod
    def _simplified_bulk_csv_signature(csv_path: str) -> dict[str, int | str]:
        resolved_path = Path(csv_path).expanduser().resolve()
        stat_result = resolved_path.stat()
        return {
            "csv_path": str(resolved_path),
            "csv_size": int(stat_result.st_size),
            "csv_mtime_ns": int(stat_result.st_mtime_ns),
        }

    def _load_simplified_bulk_checkpoint(self, csv_path: str) -> dict | None:
        checkpoint_path = self._simplified_bulk_checkpoint_path(csv_path)
        if not checkpoint_path.exists():
            return None

        try:
            with open(checkpoint_path, "r", encoding="utf-8") as file_handle:
                checkpoint = json.load(file_handle)
        except Exception as error:  # pylint: disable=broad-except
            self._append_log(f"Could not read simplified bulk checkpoint: {error}")
            return None

        try:
            signature = self._simplified_bulk_csv_signature(csv_path)
        except OSError:
            return None

        if (
            checkpoint.get("csv_path") != signature["csv_path"]
            or int(checkpoint.get("csv_size", -1)) != signature["csv_size"]
            or int(checkpoint.get("csv_mtime_ns", -1)) != signature["csv_mtime_ns"]
        ):
            return None
        return checkpoint

    def _save_simplified_bulk_checkpoint(self, checkpoint: dict) -> None:
        checkpoint_path = self._simplified_bulk_checkpoint_path(str(checkpoint["csv_path"]))
        checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = checkpoint_path.with_suffix(".tmp")
        with open(temp_path, "w", encoding="utf-8") as file_handle:
            json.dump(checkpoint, file_handle, indent=2)
        os.replace(temp_path, checkpoint_path)

    def _delete_simplified_bulk_checkpoint(self, csv_path: str) -> None:
        checkpoint_path = self._simplified_bulk_checkpoint_path(csv_path)
        if checkpoint_path.exists():
            checkpoint_path.unlink()

    @staticmethod
    def _checkpoint_phase_label(phase: str) -> str:
        if phase == "dry_run":
            return "Dry run"
        if phase == "awaiting_confirmation":
            return "Dry run complete"
        if phase == "copy":
            return "Actual copy"
        return phase.replace("_", " ").title()

    @staticmethod
    def _report_rows_from_checkpoint(raw_rows: list[dict]) -> list[SimplifiedBulkCopyReportRow]:
        return [
            SimplifiedBulkCopyReportRow(
                row_label=str(row.get("row_label", "")),
                source_uri=str(row.get("source_uri", "")),
                destination_uri=str(row.get("destination_uri", "")),
                status=str(row.get("status", "")),
                message=str(row.get("message", "")),
                destination_folder_uri=str(row.get("destination_folder_uri", "")),
                destination_folder_status=str(row.get("destination_folder_status", "")),
            )
            for row in raw_rows
        ]

    def _sync_simplified_bulk_report(self, checkpoint: dict, report_kind: str) -> Path:
        report_path_value = checkpoint.get(f"{report_kind}_report_path", "")
        report_path = Path(report_path_value) if report_path_value else self._simplified_bulk_report_path(report_kind)
        checkpoint[f"{report_kind}_report_path"] = str(report_path)
        report_rows = self._report_rows_from_checkpoint(checkpoint.get(f"{report_kind}_rows", []))
        self._write_simplified_bulk_report(report_rows, report_kind, report_path=report_path)
        return report_path

    @staticmethod
    def _summarize_simplified_bulk_rows(raw_rows: list[dict]) -> tuple[int, int, int, int]:
        ready_count = sum(1 for row in raw_rows if str(row.get("status", "")) == "ready")
        overwrite_count = sum(1 for row in raw_rows if str(row.get("status", "")) == "overwrite_warning")
        internal_conflict_count = sum(1 for row in raw_rows if str(row.get("status", "")) == "internal_conflict")
        error_count = sum(1 for row in raw_rows if str(row.get("status", "")) == "failed")
        return ready_count, overwrite_count, internal_conflict_count, error_count

    def _format_simplified_bulk_resume_summary(self, checkpoint: dict) -> str:
        phase_label = self._checkpoint_phase_label(str(checkpoint.get("phase", "dry_run")))
        next_index = int(checkpoint.get("next_index", 0))
        total_rows = int(checkpoint.get("total_rows", 0))
        paused_text = " Paused." if checkpoint.get("paused") else ""

        if checkpoint.get("phase") in {"dry_run", "awaiting_confirmation"}:
            ready_count, overwrite_count, internal_conflict_count, error_count = self._summarize_simplified_bulk_rows(
                checkpoint.get("dry_run_rows", [])
            )
            return (
                f"Resume available: {phase_label} {next_index}/{total_rows} row(s) processed."
                f" Ready {ready_count}, internal conflicts {internal_conflict_count},"
                f" overwrite warnings {overwrite_count}, errors {error_count}.{paused_text}"
            )

        success_count = sum(1 for row in checkpoint.get("copy_rows", []) if str(row.get("status", "")) == "success")
        failure_count = sum(1 for row in checkpoint.get("copy_rows", []) if str(row.get("status", "")) == "failed")
        return (
            f"Resume available: {phase_label} {next_index}/{total_rows} row(s) processed."
            f" Successes {success_count}, failures {failure_count}.{paused_text}"
        )

    def _build_new_simplified_bulk_checkpoint(self, csv_path: str, total_rows: int) -> dict:
        signature = self._simplified_bulk_csv_signature(csv_path)
        checkpoint = {
            "version": 1,
            "csv_path": signature["csv_path"],
            "csv_size": signature["csv_size"],
            "csv_mtime_ns": signature["csv_mtime_ns"],
            "phase": "dry_run",
            "total_rows": total_rows,
            "next_index": 0,
            "paused": False,
            "overwrite_mode": "",
            "copy_selection": "",
            "overwrite_next_index": 0,
            "dry_run_report_path": str(self._simplified_bulk_report_path("dry_run")),
            "copy_report_path": str(self._simplified_bulk_report_path("result")),
            "dry_run_rows": [],
            "copy_rows": [],
        }
        self._save_simplified_bulk_checkpoint(checkpoint)
        return checkpoint

    def _build_simplified_bulk_copy_plan(
        self,
        copy_items: list[tuple[str, ResolvedS3Paths]],
        checkpoint: dict,
    ) -> list[tuple[int, str, ResolvedS3Paths, str]]:
        selection = str(checkpoint.get("copy_selection", "")).strip()
        if selection == "all_rows":
            dry_run_rows = checkpoint.get("dry_run_rows", [])
            return [
                (
                    index,
                    item_label,
                    item_paths,
                    str(dry_run_rows[index].get("status", "unvalidated")) if index < len(dry_run_rows) else "unvalidated",
                )
                for index, (item_label, item_paths) in enumerate(copy_items)
            ]
        if not selection:
            overwrite_mode = str(checkpoint.get("overwrite_mode", "")).strip()
            selection = "include_overwrites" if overwrite_mode == "overwrite_all" else "ready_only"

        if selection == "include_internal_conflicts":
            allowed_statuses = {"ready", "internal_conflict"}
        elif selection == "include_overwrites":
            allowed_statuses = {"ready", "overwrite_warning"}
        else:
            allowed_statuses = {"ready"}
        dry_run_rows = checkpoint.get("dry_run_rows", [])
        planned_items: list[tuple[int, str, ResolvedS3Paths, str]] = []
        for index, (item, row) in enumerate(zip(copy_items, dry_run_rows)):
            row_status = str(row.get("status", ""))
            if row_status in allowed_statuses:
                item_label, item_paths = item
                planned_items.append((index, item_label, item_paths, row_status))
        return planned_items

    @staticmethod
    def _find_pending_overwrite_entries(
        copy_items: list[tuple[int, str, ResolvedS3Paths, str]],
        checkpoint: dict,
    ) -> list[tuple[int, int, str, ResolvedS3Paths, str]]:
        pending_entries: list[tuple[int, int, str, ResolvedS3Paths, str]] = []
        copy_rows = checkpoint.get("copy_rows", [])
        for row_index, item in enumerate(copy_items):
            if row_index >= len(copy_rows):
                break
            if str(copy_rows[row_index].get("status", "")) == "overwrite_pending":
                original_index, item_label, item_paths, row_status = item
                pending_entries.append((row_index, original_index, item_label, item_paths, row_status))
        return pending_entries

    @staticmethod
    def _find_duplicate_destination_indices(copy_items: list[tuple[str, ResolvedS3Paths]]) -> set[int]:
        destination_indices: dict[tuple[str, str], list[int]] = {}
        for index, (_item_label, item_paths) in enumerate(copy_items):
            destination_key = (item_paths.dest_bucket, item_paths.dest_key)
            destination_indices.setdefault(destination_key, []).append(index)

        duplicate_indices: set[int] = set()
        for indices in destination_indices.values():
            if len(indices) > 1:
                duplicate_indices.update(indices)
        return duplicate_indices

    def _prompt_simplified_bulk_copy_overwrite_action(
        self,
        report_path: Path,
        success_count: int,
        pending_count: int,
        failure_count: int,
    ) -> str:
        result = {"value": "cancel"}
        dialog = tk.Toplevel(self.root)
        dialog.title("Simplified Bulk Copy Overwrite Review")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        body = ttk.Frame(dialog, padding=16)
        body.pack(fill="both", expand=True)

        detail_lines = [
            f"Copied successfully: {success_count}",
            f"Would overwrite existing destination files: {pending_count}",
            f"Failed: {failure_count}",
            "",
            "Open the report to review the overwrite candidates.",
            "You can overwrite those files now, finish without overwriting them, or cancel and come back later.",
        ]

        ttk.Label(body, text="Main copy pass finished.").pack(anchor="w")
        ttk.Label(body, text="\n".join(detail_lines), justify="left").pack(anchor="w", pady=(8, 12))
        ttk.Label(body, text=f"Report: {report_path}", justify="left", wraplength=640).pack(anchor="w")

        button_row = ttk.Frame(body)
        button_row.pack(anchor="e", pady=(14, 0))

        ttk.Button(
            button_row,
            text="Open Report",
            command=lambda: self._open_report_file(report_path),
        ).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(
            button_row,
            text="Overwrite These Files",
            command=lambda: (result.__setitem__("value", "overwrite_all_pending"), dialog.destroy()),
        ).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(
            button_row,
            text="Finish Without Overwrites",
            command=lambda: (result.__setitem__("value", "finish_without_overwrites"), dialog.destroy()),
        ).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(
            button_row,
            text="Cancel",
            command=lambda: dialog.destroy(),
        ).grid(row=0, column=3)

        self._present_modal_dialog(dialog)
        dialog.wait_window()
        return result["value"]

    def _prompt_simplified_bulk_resume_action(self, checkpoint: dict) -> str:
        phase_label = self._checkpoint_phase_label(str(checkpoint.get("phase", "dry_run")))
        next_index = int(checkpoint.get("next_index", 0))
        total_rows = int(checkpoint.get("total_rows", 0))
        message = (
            "A previous Simplified Bulk Copy session was found for this CSV.\n\n"
            f"Phase: {phase_label}\n"
            f"Progress: {next_index}/{total_rows} row(s)\n\n"
            "Yes = Resume\n"
            "No = Restart from the beginning\n"
            "Cancel = Stop"
        )
        response = messagebox.askyesnocancel("Resume Simplified Bulk Copy", message, parent=self.root)
        if response is True:
            return "resume"
        if response is False:
            return "restart"
        return "cancel"

    def _set_pause_state(self, requested: bool, active: bool | None = None) -> None:
        self._pause_requested = requested
        if active is not None:
            self._pause_active = active
        self._update_pause_button_state()

    def _update_pause_button_state(self) -> None:
        if self._running and self._is_simplified_bulk_mode():
            self.pause_button.configure(
                state="normal",
                text="Resume" if (self._pause_requested or self._pause_active) else "Pause",
            )
        else:
            self.pause_button.configure(state="disabled", text="Pause")

    def _on_pause_resume_clicked(self) -> None:
        if not self._running or not self._is_simplified_bulk_mode():
            return

        if self._pause_requested or self._pause_active:
            self._set_pause_state(False, active=False)
            self._append_log("Resume requested for Simplified Bulk Copy.")
            return

        self._set_pause_state(True)
        self._append_log("Pause requested. The current row will finish before the session pauses.")

    def _wait_if_simplified_bulk_paused(self, checkpoint: dict, phase_label: str) -> None:
        if not self._pause_requested:
            return

        if not checkpoint.get("paused"):
            checkpoint["paused"] = True
            self._save_simplified_bulk_checkpoint(checkpoint)
            self._enqueue_ui(self._append_log, f"{phase_label} paused. Click Resume to continue.")

        self._enqueue_ui(self._set_pause_state, True, True)
        while self._pause_requested and not self._closing:
            time.sleep(0.2)

        checkpoint["paused"] = False
        self._save_simplified_bulk_checkpoint(checkpoint)
        self._enqueue_ui(self._set_pause_state, False, False)
        self._enqueue_ui(self._append_log, f"{phase_label} resumed.")

    def _is_direct_upload_mode(self) -> bool:
        selected_tab = self.mode_notebook.select()
        return selected_tab == str(self.direct_mode_frame)

    def _is_rename_mode(self) -> bool:
        selected_tab = self.mode_notebook.select()
        return selected_tab == str(self.rename_mode_frame)

    def _is_simplified_bulk_mode(self) -> bool:
        selected_tab = self.mode_notebook.select()
        return selected_tab == str(self.simplified_bulk_mode_frame)

    def _is_inventory_mode(self) -> bool:
        selected_tab = self.mode_notebook.select()
        return selected_tab == str(self.inventory_mode_frame)

    def _is_folder_copy_mode(self) -> bool:
        if self.folder_copy_mode_frame is None:
            return False
        selected_tab = self.mode_notebook.select()
        return selected_tab == str(self.folder_copy_mode_frame)

    def _is_bulk_folder_copy_mode(self) -> bool:
        if self.bulk_folder_copy_mode_frame is None:
            return False
        selected_tab = self.mode_notebook.select()
        return selected_tab == str(self.bulk_folder_copy_mode_frame)

    def _download_template_for_mode(self, mode: str) -> None:
        if mode == "direct_upload":
            columns = BulkCopyDialog.DIRECT_UPLOAD_COLUMNS
            template_rows = [
                ("1", "/Users/your.name/Downloads/source_video_1.mp4", "folder1/folder2", "new_video_name.mp4"),
                ("2", "/Users/your.name/Downloads/source_video_2.mp4", "folder1/folder2", "new_video_name_2.mp4"),
            ]
            initial_file = "bulk_upload_title_mapping.csv"
        else:
            columns = BulkCopyDialog.S3_COPY_COLUMNS
            template_rows = [
                ("Title 1", "current_video_name.mp4", "", "folder1/folder2", "new_video_name.mp4", ""),
                ("Title 2", "current_video_name_2.mp4", "current_video_name_2.vtt", "folder1/folder2", "new_video_name_2.mp4", "new_video_name_2.vtt"),
            ]
            initial_file = "bulk_copy_template.csv"

        save_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="Save Template",
            defaultextension=".csv",
            initialfile=initial_file,
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*")),
        )
        if not save_path:
            return

        headers = [heading for _, heading, _ in columns]
        try:
            with open(save_path, "w", encoding="utf-8", newline="") as file_handle:
                writer = csv.writer(file_handle)
                writer.writerow(headers)
                writer.writerows(template_rows)
            messagebox.showinfo("Template Saved", f"Template saved to:\n{save_path}", parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            messagebox.showerror("Template Save Failed", str(error), parent=self.root)

    def download_bulk_copy_template(self) -> None:
        self._download_template_for_mode("s3_copy")

    def download_bulk_upload_template(self) -> None:
        self._download_template_for_mode("direct_upload")

    def _on_settings_saved(self, new_config: AppConfig, session_credentials: AwsCredentials | None) -> None:
        self.config = new_config
        self.use_session_only_credentials = self.config.credential_mode == "session"
        self.session_credentials = session_credentials if self.use_session_only_credentials else None
        if self.use_session_only_credentials:
            self.keychain_credentials = None
            self._keychain_credentials_loaded = False
        else:
            try:
                self.keychain_credentials = load_credentials()
                self._keychain_credentials_loaded = True
            except RuntimeError as error:
                self._append_log(str(error))
                self.keychain_credentials = None
                self._keychain_credentials_loaded = False

        if self.use_session_only_credentials:
            if self.session_credentials:
                self._append_log("Settings updated. Using session-only credentials (not saved to Keychain).")
            else:
                self._append_log("Settings updated. Session-only mode is active with no credentials saved.")
        else:
            self._append_log("Settings updated. Using Keychain/default AWS credential chain.")
        self._refresh_preview()

    def _load_keychain_credentials(self, refresh: bool = False) -> AwsCredentials | None:
        if self.use_session_only_credentials:
            return None
        if self._keychain_credentials_loaded and not refresh:
            return self.keychain_credentials

        credentials = load_credentials()
        self.keychain_credentials = credentials
        self._keychain_credentials_loaded = True
        return credentials

    def _get_active_credentials(self) -> AwsCredentials | None:
        if self.use_session_only_credentials:
            return self.session_credentials
        return self._load_keychain_credentials()

    @staticmethod
    def _parse_s3_folder_uri(value: str) -> tuple[str, str, str]:
        bucket, key = S3CopyApp._parse_s3_uri(value)
        normalized_prefix = sanitize_folder_path(key)
        if not normalized_prefix:
            raise ValueError("Folder URI must include a prefix after the bucket name.")
        normalized_prefix = normalized_prefix.rstrip("/") + "/"
        return bucket, normalized_prefix, f"s3://{bucket}/{normalized_prefix}"

    @staticmethod
    def _parse_s3_inventory_uri(value: str) -> tuple[str, str, str]:
        cleaned = value.strip()
        if not cleaned:
            raise ValueError("S3 URI cannot be blank.")
        if not cleaned.lower().startswith("s3://"):
            raise ValueError(f"Invalid S3 URI: {cleaned}")

        remainder = cleaned[5:]
        bucket, separator, key = remainder.partition("/")
        bucket = bucket.strip()
        if not bucket:
            raise ValueError(f"Invalid S3 URI bucket: {cleaned}")

        if not separator or not key.strip():
            return bucket, "", f"s3://{bucket}/"

        normalized_prefix = sanitize_folder_path(key)
        if not normalized_prefix:
            return bucket, "", f"s3://{bucket}/"
        normalized_prefix = normalized_prefix.rstrip("/") + "/"
        return bucket, normalized_prefix, f"s3://{bucket}/{normalized_prefix}"

    @staticmethod
    def _inventory_report_path() -> Path:
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.exists():
            base_dir = downloads_dir
        else:
            base_dir = Path.home()
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        return base_dir / f"{APP_FILE_SLUG}_inventory_{timestamp}.csv"

    def _write_inventory_report(
        self,
        inventory_uri: str,
        listed_objects: list[S3ListedObject],
        report_path: Path | None = None,
    ) -> Path:
        if report_path is None:
            report_path = self._inventory_report_path()
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w", encoding="utf-8", newline="") as file_handle:
            writer = csv.writer(file_handle)
            writer.writerow(["inventory_uri", inventory_uri])
            writer.writerow([])
            writer.writerow(["bucket", "key", "size_bytes", "last_modified", "s3_uri"])
            for item in listed_objects:
                writer.writerow(
                    [
                        item.bucket,
                        item.key,
                        item.size_bytes,
                        item.last_modified,
                        f"s3://{item.bucket}/{item.key}",
                    ]
                )
        return report_path

    def _build_folder_copy_items(
        self,
        s3_client,
        source_folder_uri: str,
        destination_folder_uri: str,
    ) -> tuple[list[tuple[str, ResolvedS3Paths]], FolderCopyPreview]:
        source_bucket, source_prefix, normalized_source_uri = self._parse_s3_folder_uri(source_folder_uri)
        dest_bucket, dest_prefix, normalized_dest_uri = self._parse_s3_folder_uri(destination_folder_uri)

        if source_bucket == dest_bucket and source_prefix == dest_prefix:
            raise UserVisibleError("Source folder and destination folder resolve to the same S3 location.")

        self._enqueue_ui(self._append_log, f"Scanning source folder: {normalized_source_uri}")
        source_objects = list_objects_under_prefix(
            s3_client,
            source_bucket,
            source_prefix,
            progress_callback=lambda msg: self._enqueue_ui(self._append_log, f"Folder scan: {msg}"),
        )
        if not source_objects:
            raise UserVisibleError("No objects were found under the source folder.")

        copy_items: list[tuple[str, ResolvedS3Paths]] = []
        first_source_uri = ""
        first_destination_uri = ""

        for index, source_object in enumerate(source_objects, start=1):
            relative_key = source_object.key[len(source_prefix) :]
            if not relative_key:
                continue

            dest_key = join_key_parts(dest_prefix, relative_key)
            paths = ResolvedS3Paths(
                source_bucket=source_bucket,
                source_key=source_object.key,
                dest_bucket=dest_bucket,
                dest_key=dest_key,
            )
            copy_items.append((f"Folder item {index}", paths))

            if not first_source_uri:
                first_source_uri = paths.source_uri
                first_destination_uri = paths.dest_uri

        if not copy_items:
            raise UserVisibleError("No file objects were found under the source folder.")

        preview = FolderCopyPreview(
            object_count=len(copy_items),
            first_source_uri=first_source_uri,
            first_destination_uri=first_destination_uri,
        )
        return copy_items, preview

    def _build_bulk_folder_copy_items(
        self,
        s3_client,
        folder_jobs: list[tuple[str, str]],
    ) -> tuple[list[tuple[str, ResolvedS3Paths]], FolderCopyPreview]:
        combined_items: list[tuple[str, ResolvedS3Paths]] = []
        first_source_uri = ""
        first_destination_uri = ""

        for job_index, (source_folder_uri, destination_folder_uri) in enumerate(folder_jobs, start=1):
            self._enqueue_ui(
                self._append_log,
                f"Scanning folder pair {job_index}/{len(folder_jobs)}: {source_folder_uri} -> {destination_folder_uri}",
            )
            job_items, preview = self._build_folder_copy_items(s3_client, source_folder_uri, destination_folder_uri)
            for item_index, (_old_label, item_paths) in enumerate(job_items, start=1):
                combined_items.append((f"Folder {job_index} item {item_index}", item_paths))
            if not first_source_uri:
                first_source_uri = preview.first_source_uri
                first_destination_uri = preview.first_destination_uri

        if not combined_items:
            raise UserVisibleError("No file objects were found under the provided folder pairs.")

        return combined_items, FolderCopyPreview(
            object_count=len(combined_items),
            first_source_uri=first_source_uri,
            first_destination_uri=first_destination_uri,
            folder_pair_count=len(folder_jobs),
        )

    def _prepare_copy_items(
        self,
        user_input: UserInput,
        label_prefix: str,
    ) -> tuple[list[str], list[tuple[str, ResolvedS3Paths]]]:
        validation_errors = validate_user_input(self.config, user_input)
        primary_paths = build_paths(self.config, user_input)
        validation_errors.extend(validate_paths_not_identical(primary_paths))

        caption_paths = self._build_caption_paths(user_input)
        if caption_paths:
            validation_errors.extend(validate_paths_not_identical(caption_paths))
            if (
                primary_paths.source_bucket == caption_paths.source_bucket
                and primary_paths.source_key == caption_paths.source_key
            ):
                validation_errors.append(
                    "Current Caption Name resolves to the same source object as Current File Name."
                )
            if (
                primary_paths.dest_bucket == caption_paths.dest_bucket
                and primary_paths.dest_key == caption_paths.dest_key
            ):
                validation_errors.append("Desired Caption Name resolves to the same destination as Desired Name.")

        copy_items: list[tuple[str, ResolvedS3Paths]] = [(f"{label_prefix} - Primary file", primary_paths)]
        if caption_paths:
            copy_items.append((f"{label_prefix} - Caption file", caption_paths))

        return validation_errors, copy_items

    def _prepare_direct_upload_items(
        self,
        label_prefix: str,
        local_file_path: str,
        local_caption_path: str,
        desired_move_folder: str,
        desired_name: str,
        desired_caption_name: str,
    ) -> tuple[list[str], list[DirectUploadItem]]:
        local_file_path = local_file_path.strip()
        local_caption_path = local_caption_path.strip()
        desired_move_folder = sanitize_folder_path(desired_move_folder)
        desired_name = sanitize_filename(desired_name)
        desired_caption_name = sanitize_filename(desired_caption_name)

        errors: list[str] = []
        if not local_file_path:
            errors.append("Local File is required for direct upload.")
        elif not os.path.isfile(local_file_path):
            errors.append(f"Local File not found: {local_file_path}")

        if not desired_move_folder:
            errors.append("Desired Move Folder cannot be blank.")
        if not desired_name:
            errors.append("Desired Name cannot be blank.")
        if "/" in desired_name:
            errors.append("Desired Name must be a file name only (no slashes).")
        if desired_name and "." not in desired_name.strip("."):
            errors.append("Desired Name must include a file extension (example: report.pdf).")

        if bool(local_caption_path) ^ bool(desired_caption_name):
            errors.append(
                "For optional caption upload, provide both Local Caption and Desired Caption Name, or leave both blank."
            )
        if local_caption_path and not os.path.isfile(local_caption_path):
            errors.append(f"Local Caption not found: {local_caption_path}")
        if "/" in desired_caption_name:
            errors.append("Desired Caption Name must be a file name only (no slashes).")
        if desired_caption_name and "." not in desired_caption_name.strip("."):
            errors.append("Desired Caption Name must include a file extension (example: trailer_en.vtt).")

        dest_bucket = self.config.dest_bucket.strip()
        if not dest_bucket:
            errors.append("Destination bucket is not configured. Open Settings.")

        items: list[DirectUploadItem] = []
        if errors:
            return errors, items

        primary_dest_key = join_key_parts(self.config.dest_prefix, desired_move_folder, desired_name)
        primary_dest_ref = S3ObjectRef(bucket=dest_bucket, key=primary_dest_key)
        items.append(
            DirectUploadItem(
                label=f"{label_prefix} - Primary file",
                local_path=local_file_path,
                destination_ref=primary_dest_ref,
                destination_uri=f"s3://{dest_bucket}/{primary_dest_key}",
            )
        )

        if local_caption_path and desired_caption_name:
            caption_dest_key = join_key_parts(self.config.dest_prefix, desired_move_folder, desired_caption_name)
            if caption_dest_key == primary_dest_key:
                errors.append("Desired Caption Name resolves to the same destination as Desired Name.")
                return errors, []

            caption_dest_ref = S3ObjectRef(bucket=dest_bucket, key=caption_dest_key)
            items.append(
                DirectUploadItem(
                    label=f"{label_prefix} - Caption file",
                    local_path=local_caption_path,
                    destination_ref=caption_dest_ref,
                    destination_uri=f"s3://{dest_bucket}/{caption_dest_key}",
                )
            )

        return errors, items

    def _prepare_rename_item(self) -> tuple[list[str], tuple[S3ObjectRef, S3ObjectRef, str, str] | None]:
        current_path = sanitize_folder_path(self.rename_current_path_var.get())
        current_name = sanitize_filename(self.rename_current_name_var.get())
        desired_name = sanitize_filename(self.rename_desired_name_var.get())

        errors: list[str] = []
        if not current_name:
            errors.append("Current File Name cannot be blank.")
        if not desired_name:
            errors.append("Desired Name cannot be blank.")

        if "/" in current_name:
            errors.append("Current File Name must be a file name only (no slashes).")
        if "/" in desired_name:
            errors.append("Desired Name must be a file name only (no slashes).")
        if desired_name and "." not in desired_name.strip("."):
            errors.append("Desired Name must include a file extension (example: report.pdf).")

        dest_bucket = self.config.dest_bucket.strip()
        if not dest_bucket:
            errors.append("Destination bucket is not configured. Open Settings.")

        source_key = join_key_parts(self.config.dest_prefix, current_path, current_name)
        dest_key = join_key_parts(self.config.dest_prefix, current_path, desired_name)
        if source_key == dest_key:
            errors.append("Current Destination Name and Desired Name resolve to the same S3 object.")

        if errors:
            return errors, None

        source_ref = S3ObjectRef(bucket=dest_bucket, key=source_key)
        dest_ref = S3ObjectRef(bucket=dest_bucket, key=dest_key)
        source_uri = f"s3://{dest_bucket}/{source_key}"
        dest_uri = f"s3://{dest_bucket}/{dest_key}"
        return errors, (source_ref, dest_ref, source_uri, dest_uri)

    @staticmethod
    def _parse_s3_uri(value: str) -> tuple[str, str]:
        cleaned = value.strip()
        if not cleaned:
            raise ValueError("S3 URI cannot be blank.")
        if not cleaned.lower().startswith("s3://"):
            raise ValueError(f"Invalid S3 URI: {cleaned}")

        remainder = cleaned[5:]
        bucket, separator, key = remainder.partition("/")
        if not bucket:
            raise ValueError(f"Invalid S3 URI bucket: {cleaned}")
        if not separator or not key:
            raise ValueError(f"Invalid S3 URI key: {cleaned}")
        return bucket.strip(), key.strip()

    def _load_simplified_bulk_copy_items(
        self,
        csv_path: str,
    ) -> tuple[list[str], list[tuple[str, ResolvedS3Paths]], SimplifiedBulkCsvPreview | None]:
        file_path = csv_path.strip()
        errors: list[str] = []
        copy_items: list[tuple[str, ResolvedS3Paths]] = []

        if not file_path:
            return ["CSV File cannot be blank."], copy_items, None
        if not os.path.isfile(file_path):
            return [f"CSV File not found: {file_path}"], copy_items, None
        if Path(file_path).suffix.lower() != ".csv":
            return ["Simplified bulk copy requires a .csv file."], copy_items, None

        with open(file_path, "r", encoding="utf-8-sig", newline="") as file_handle:
            reader = csv.DictReader(file_handle)
            if not reader.fieldnames:
                return ["No header row found in CSV file."], copy_items, None

            normalized_headers = {
                header.strip().lower().replace(" ", "_"): header
                for header in reader.fieldnames
                if header
            }
            missing_headers = [
                header_name for header_name in SIMPLIFIED_BULK_REQUIRED_COLUMNS if header_name not in normalized_headers
            ]
            if missing_headers:
                missing_text = ", ".join(missing_headers)
                return [f"CSV must include these columns: {missing_text}"], copy_items, None

            first_source_uri = ""
            first_destination_uri = ""
            populated_row_count = 0

            for csv_row_number, row in enumerate(reader, start=2):
                source_uri = str(row.get(normalized_headers["source_uri"], "")).strip()
                destination_uri = str(row.get(normalized_headers["destination_uri"], "")).strip()

                if not source_uri and not destination_uri:
                    continue

                row_label = f"CSV Row {csv_row_number}"
                populated_row_count += 1
                if not first_source_uri:
                    first_source_uri = source_uri
                    first_destination_uri = destination_uri

                if not source_uri:
                    errors.append(f"{row_label}: source_uri cannot be blank.")
                    continue
                if not destination_uri:
                    errors.append(f"{row_label}: destination_uri cannot be blank.")
                    continue

                try:
                    source_bucket, source_key = self._parse_s3_uri(source_uri)
                    dest_bucket, dest_key = self._parse_s3_uri(destination_uri)
                except ValueError as error:
                    errors.append(f"{row_label}: {error}")
                    continue

                paths = ResolvedS3Paths(
                    source_bucket=source_bucket,
                    source_key=source_key,
                    dest_bucket=dest_bucket,
                    dest_key=dest_key,
                )
                row_errors = validate_paths_not_identical(paths)
                if row_errors:
                    errors.extend([f"{row_label}: {message}" for message in row_errors])
                    continue

                copy_items.append((row_label, paths))

        if not copy_items and not errors:
            errors.append("No populated rows found in the CSV file.")

        preview = None
        if first_source_uri and first_destination_uri:
            preview = SimplifiedBulkCsvPreview(
                row_count=len(copy_items),
                first_source_uri=first_source_uri,
                first_destination_uri=first_destination_uri,
            )

        return errors, copy_items, preview

    def _update_simplified_bulk_summary(self) -> None:
        csv_path = self.simplified_bulk_csv_path_var.get().strip()
        if not csv_path:
            self.simplified_bulk_summary_var.set(
                "Load a CSV with source_uri and destination_uri columns. Resume details appear here if a prior run exists."
            )
            return

        errors, copy_items, _preview = self._load_simplified_bulk_copy_items(csv_path)
        if errors:
            self.simplified_bulk_summary_var.set(errors[0])
            return

        checkpoint = self._load_simplified_bulk_checkpoint(csv_path)
        if checkpoint:
            self.simplified_bulk_summary_var.set(self._format_simplified_bulk_resume_summary(checkpoint))
            return

        self.simplified_bulk_summary_var.set(
            f"Ready: {len(copy_items)} row(s) loaded from {Path(csv_path).name}."
        )

    def _load_bulk_folder_copy_jobs(
        self,
        csv_path: str,
    ) -> tuple[list[str], list[tuple[str, str]], FolderCopyPreview | None]:
        file_path = csv_path.strip()
        errors: list[str] = []
        folder_jobs: list[tuple[str, str]] = []

        if not file_path:
            return ["CSV File cannot be blank."], folder_jobs, None
        if not os.path.isfile(file_path):
            return [f"CSV File not found: {file_path}"], folder_jobs, None
        if Path(file_path).suffix.lower() != ".csv":
            return ["Bulk folder copy requires a .csv file."], folder_jobs, None

        with open(file_path, "r", encoding="utf-8-sig", newline="") as file_handle:
            reader = csv.DictReader(file_handle)
            if not reader.fieldnames:
                return ["No header row found in CSV file."], folder_jobs, None

            normalized_headers = {
                header.strip().lower().replace(" ", "_"): header
                for header in reader.fieldnames
                if header
            }
            missing_headers = [
                header_name for header_name in BULK_FOLDER_REQUIRED_COLUMNS if header_name not in normalized_headers
            ]
            if missing_headers:
                missing_text = ", ".join(missing_headers)
                return [f"CSV must include these columns: {missing_text}"], folder_jobs, None

            first_source_uri = ""
            first_destination_uri = ""

            for csv_row_number, row in enumerate(reader, start=2):
                source_folder_uri = str(row.get(normalized_headers["source_folder_uri"], "")).strip()
                destination_folder_uri = str(row.get(normalized_headers["destination_folder_uri"], "")).strip()

                if not source_folder_uri and not destination_folder_uri:
                    continue

                row_label = f"CSV Row {csv_row_number}"
                if not source_folder_uri:
                    errors.append(f"{row_label}: source_folder_uri cannot be blank.")
                    continue
                if not destination_folder_uri:
                    errors.append(f"{row_label}: destination_folder_uri cannot be blank.")
                    continue

                try:
                    _source_bucket, _source_prefix, normalized_source_uri = self._parse_s3_folder_uri(source_folder_uri)
                    _dest_bucket, _dest_prefix, normalized_dest_uri = self._parse_s3_folder_uri(destination_folder_uri)
                except ValueError as error:
                    errors.append(f"{row_label}: {error}")
                    continue

                if normalized_source_uri == normalized_dest_uri:
                    errors.append(f"{row_label}: Source folder and destination folder resolve to the same S3 location.")
                    continue

                if not first_source_uri:
                    first_source_uri = normalized_source_uri
                    first_destination_uri = normalized_dest_uri

                folder_jobs.append((normalized_source_uri, normalized_dest_uri))

        if not folder_jobs and not errors:
            errors.append("No populated rows found in the CSV file.")

        preview = None
        if first_source_uri and first_destination_uri:
            preview = FolderCopyPreview(
                object_count=0,
                first_source_uri=first_source_uri,
                first_destination_uri=first_destination_uri,
                folder_pair_count=len(folder_jobs),
            )

        return errors, folder_jobs, preview

    def _update_bulk_folder_summary(self) -> None:
        csv_path = self.bulk_folder_csv_path_var.get().strip()
        if not csv_path:
            self.bulk_folder_summary_var.set(
                "Load a CSV with source_folder_uri and destination_folder_uri columns to copy multiple folder pairs."
            )
            return

        errors, folder_jobs, _preview = self._load_bulk_folder_copy_jobs(csv_path)
        if errors:
            self.bulk_folder_summary_var.set(errors[0])
            return

        self.bulk_folder_summary_var.set(
            f"Ready: {len(folder_jobs)} folder pair(s) loaded from {Path(csv_path).name}."
        )

    @staticmethod
    def _build_s3_ref_from_paths(paths: ResolvedS3Paths) -> tuple[S3ObjectRef, S3ObjectRef]:
        source_ref = S3ObjectRef(bucket=paths.source_bucket, key=paths.source_key)
        dest_ref = S3ObjectRef(bucket=paths.dest_bucket, key=paths.dest_key)
        return source_ref, dest_ref

    @staticmethod
    def _destination_folder_details(paths: ResolvedS3Paths) -> tuple[str, str]:
        parent_prefix = sanitize_folder_path(paths.dest_key.rpartition("/")[0])
        if not parent_prefix:
            return f"s3://{paths.dest_bucket}/", ""
        normalized_prefix = parent_prefix.rstrip("/") + "/"
        return f"s3://{paths.dest_bucket}/{normalized_prefix}", normalized_prefix

    def _resolve_destination_folder_status(
        self,
        s3_client,
        paths: ResolvedS3Paths,
        cache: dict[tuple[str, str], bool],
    ) -> tuple[str, str, str]:
        folder_uri, folder_prefix = self._destination_folder_details(paths)
        if not folder_prefix:
            return folder_uri, "bucket_root_exists", "Destination bucket root already exists."

        cache_key = (paths.dest_bucket, folder_prefix)
        if cache_key not in cache:
            cache[cache_key] = prefix_exists(s3_client, paths.dest_bucket, folder_prefix)

        if cache[cache_key]:
            return folder_uri, "exists", "Destination folder path already exists."
        return folder_uri, "will_create", "Destination folder path will be created by this copy."

    def _build_aws_cli_dry_run_environment(self, credentials: AwsCredentials | None) -> dict[str, str]:
        env = os.environ.copy()
        if credentials:
            env["AWS_ACCESS_KEY_ID"] = credentials.access_key_id
            env["AWS_SECRET_ACCESS_KEY"] = credentials.secret_access_key
            if credentials.session_token:
                env["AWS_SESSION_TOKEN"] = credentials.session_token
            else:
                env.pop("AWS_SESSION_TOKEN", None)
        if self.config.aws_region.strip():
            env["AWS_DEFAULT_REGION"] = self.config.aws_region.strip()
        return env

    def _run_aws_cli_dry_run(
        self,
        aws_cli_path: str,
        credentials: AwsCredentials | None,
        source_uri: str,
        dest_uri: str,
    ) -> tuple[bool, str]:
        command = [aws_cli_path, "s3", "cp", source_uri, dest_uri, "--dryrun"]
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            env=self._build_aws_cli_dry_run_environment(credentials),
            check=False,
        )
        output = (completed.stdout or completed.stderr or "").strip()
        if completed.returncode != 0:
            return False, output or "AWS CLI dry run failed."
        return True, output or "AWS CLI dry run completed."

    @staticmethod
    def _find_aws_cli_path() -> str | None:
        aws_cli_path = shutil.which("aws")
        if aws_cli_path:
            return aws_cli_path

        for candidate_path in ("/opt/homebrew/bin/aws", "/usr/local/bin/aws"):
            if os.path.isfile(candidate_path) and os.access(candidate_path, os.X_OK):
                return candidate_path
        return None

    @staticmethod
    def _simplified_bulk_report_path(report_kind: str) -> Path:
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.exists():
            base_dir = downloads_dir
        else:
            base_dir = Path.home()
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        return base_dir / f"{APP_FILE_SLUG}_simplified_bulk_{report_kind}_{timestamp}.csv"

    def _write_simplified_bulk_report(
        self,
        report_rows: list[SimplifiedBulkCopyReportRow],
        report_kind: str,
        report_path: Path | None = None,
    ) -> Path:
        if report_path is None:
            report_path = self._simplified_bulk_report_path(report_kind)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w", encoding="utf-8", newline="") as file_handle:
            writer = csv.writer(file_handle)
            writer.writerow(
                [
                    "row_label",
                    "source_uri",
                    "destination_uri",
                    "destination_folder_uri",
                    "destination_folder_status",
                    "status",
                    "message",
                ]
            )
            for row in report_rows:
                writer.writerow(
                    [
                        row.row_label,
                        row.source_uri,
                        row.destination_uri,
                        row.destination_folder_uri,
                        row.destination_folder_status,
                        row.status,
                        row.message,
                    ]
                )
        return report_path

    def _open_report_file(self, report_path: Path) -> None:
        try:
            if IS_WINDOWS:
                os.startfile(str(report_path))  # type: ignore[attr-defined]
            elif IS_MACOS:
                subprocess.Popen(["open", str(report_path)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                webbrowser.open(report_path.as_uri())
        except Exception as error:  # pylint: disable=broad-except
            self._append_log(f"Could not open report file: {error}")
            messagebox.showerror("Open Report Failed", f"Could not open report file: {error}", parent=self.root)

    def _prompt_simplified_bulk_preflight_action(
        self,
        dry_run_report_path: Path,
        total_rows: int,
        ready_count: int,
        overwrite_count: int,
        internal_conflict_count: int,
        error_count: int,
    ) -> str:
        result = {"value": "cancel"}
        dialog = tk.Toplevel(self.root)
        dialog.title("Simplified Bulk Copy Dry Run")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        body = ttk.Frame(dialog, padding=16)
        body.pack(fill="both", expand=True)

        if error_count or overwrite_count or internal_conflict_count:
            header_text = "Dry run found rows that need review."
            detail_lines = [
                f"Rows checked: {total_rows}",
                f"Ready: {ready_count}",
                f"Internal conflicts: {internal_conflict_count}",
                f"Overwrite warnings: {overwrite_count}",
                f"Errors: {error_count}",
                "",
                "Choose the row scope you want to run.",
            ]
        else:
            header_text = "Dry run completed with no blocking issues."
            detail_lines = [
                f"Rows checked: {total_rows}",
                f"Ready: {ready_count}",
                f"Internal conflicts: {internal_conflict_count}",
                f"Overwrite warnings: {overwrite_count}",
                f"Errors: {error_count}",
                "",
                "Open the report if you want a record, then choose a copy scope or cancel.",
            ]

        ttk.Label(body, text=header_text).pack(anchor="w")
        ttk.Label(body, text="\n".join(detail_lines), justify="left").pack(anchor="w", pady=(8, 12))
        ttk.Label(body, text=f"Report: {dry_run_report_path}", justify="left", wraplength=640).pack(anchor="w")

        button_row = ttk.Frame(body)
        button_row.pack(anchor="e", pady=(14, 0))

        ttk.Button(
            button_row,
            text="Open Report",
            command=lambda: self._open_report_file(dry_run_report_path),
        ).grid(row=0, column=0, padx=(0, 8))

        if ready_count:
            ttk.Button(
                button_row,
                text="Copy Safe Rows Only",
                command=lambda: (result.__setitem__("value", "ready_only"), dialog.destroy()),
            ).grid(row=0, column=1, padx=(0, 8))

        if ready_count or internal_conflict_count:
            ttk.Button(
                button_row,
                text="Copy Safe and Internal Conflict Rows",
                command=lambda: (result.__setitem__("value", "include_internal_conflicts"), dialog.destroy()),
            ).grid(row=0, column=2, padx=(0, 8))

        if ready_count or overwrite_count:
            ttk.Button(
                button_row,
                text="Copy Safe and Overwrite Rows",
                command=lambda: (result.__setitem__("value", "include_overwrites"), dialog.destroy()),
            ).grid(row=0, column=3, padx=(0, 8))

        if total_rows:
            ttk.Button(
                button_row,
                text="Copy All Rows",
                command=lambda: (result.__setitem__("value", "all_rows"), dialog.destroy()),
            ).grid(row=0, column=4, padx=(0, 8))

        ttk.Button(
            button_row,
            text="Cancel",
            command=lambda: dialog.destroy(),
        ).grid(row=0, column=5)

        self._present_modal_dialog(dialog)
        dialog.wait_window()
        return result["value"]

    def _on_simplified_bulk_copy_clicked(self) -> None:
        csv_path = self.simplified_bulk_csv_path_var.get().strip()
        errors, copy_items, _preview = self._load_simplified_bulk_copy_items(csv_path)
        if errors:
            messagebox.showerror("Simplified Bulk Copy Validation", "\n".join(errors), parent=self.root)
            self._append_log(f"Simplified bulk copy validation failed: {' | '.join(errors)}")
            return

        if not copy_items:
            messagebox.showerror("Simplified Bulk Copy", "No populated rows found in the CSV file.", parent=self.root)
            return

        checkpoint = self._load_simplified_bulk_checkpoint(csv_path)
        checkpoint_loaded = checkpoint is not None
        if checkpoint:
            resume_action = self._prompt_simplified_bulk_resume_action(checkpoint)
            if resume_action == "cancel":
                self._append_log("Simplified bulk copy cancelled before session resume.")
                return
            if resume_action == "restart":
                self._delete_simplified_bulk_checkpoint(csv_path)
                checkpoint = self._build_new_simplified_bulk_checkpoint(csv_path, len(copy_items))
                self._append_log("Restarting Simplified Bulk Copy from the beginning.")
            else:
                self._append_log(
                    f"Resuming Simplified Bulk Copy from {self._checkpoint_phase_label(str(checkpoint.get('phase', 'dry_run'))).lower()}."
                )
        else:
            checkpoint = self._build_new_simplified_bulk_checkpoint(csv_path, len(copy_items))

        self._set_pause_state(False, active=False)

        if not SIMPLIFIED_BULK_REQUIRE_DRY_RUN and not checkpoint_loaded:
            confirm_message = (
                f"Run actual simplified bulk copy for {len(copy_items)} row(s) now?\n\n"
                "Dry run is optional in this app. Use Run Dry Run first if you want a report before copying."
            )
            if not messagebox.askokcancel("Confirm Simplified Bulk Copy", confirm_message, parent=self.root):
                self._append_log("Simplified bulk copy cancelled before execution.")
                self._delete_simplified_bulk_checkpoint(csv_path)
                return

            checkpoint["phase"] = "copy"
            checkpoint["overwrite_mode"] = "collect_overwrites"
            checkpoint["copy_selection"] = "all_rows"
            checkpoint["next_index"] = 0
            checkpoint["overwrite_next_index"] = 0
            checkpoint["paused"] = False
            checkpoint["copy_rows"] = []
            self._save_simplified_bulk_checkpoint(checkpoint)
            self._set_running(True)
            self._append_log(f"Starting simplified bulk copy without dry run for {len(copy_items)} row(s).")
            planned_items = [
                (index, item_label, item_paths, "unvalidated")
                for index, (item_label, item_paths) in enumerate(copy_items)
            ]
            threading.Thread(
                target=self._simplified_bulk_copy_worker,
                args=(planned_items, checkpoint),
                daemon=True,
            ).start()
            return

        self._set_running(True)
        if checkpoint.get("phase") in {"copy", "copy_overwrite_review", "copy_overwrite"}:
            planned_items = self._build_simplified_bulk_copy_plan(copy_items, checkpoint)
            self._append_log(
                f"Resuming simplified bulk copy from {self._checkpoint_phase_label(str(checkpoint.get('phase', 'copy'))).lower()}."
            )
            threading.Thread(
                target=self._simplified_bulk_copy_worker,
                args=(planned_items, checkpoint),
                daemon=True,
            ).start()
            return

        if checkpoint.get("phase") == "awaiting_confirmation":
            self._append_log("Reopening previous dry run results for review.")
        else:
            self._append_log(f"Starting required simplified bulk dry run for {len(copy_items)} row(s).")
        threading.Thread(
            target=self._simplified_bulk_preflight_and_copy_worker,
            args=(copy_items, checkpoint),
            daemon=True,
        ).start()

    def _on_simplified_bulk_dry_run_clicked(self) -> None:
        if self._running:
            return

        csv_path = self.simplified_bulk_csv_path_var.get().strip()
        errors, copy_items, _preview = self._load_simplified_bulk_copy_items(csv_path)
        if errors:
            messagebox.showerror("Simplified Bulk Copy Validation", "\n".join(errors), parent=self.root)
            self._append_log(f"Simplified bulk copy validation failed: {' | '.join(errors)}")
            return

        if not copy_items:
            messagebox.showerror("Simplified Bulk Copy", "No populated rows found in the CSV file.", parent=self.root)
            return

        checkpoint = self._load_simplified_bulk_checkpoint(csv_path)
        if checkpoint and checkpoint.get("phase") == "copy":
            resume_action = self._prompt_simplified_bulk_resume_action(checkpoint)
            if resume_action == "cancel":
                self._append_log("Simplified bulk dry run cancelled before session review.")
                return
            if resume_action == "restart":
                self._delete_simplified_bulk_checkpoint(csv_path)
                checkpoint = None
            else:
                self._append_log("Existing copy session found. Reopening it through the dry-run flow is not supported.")
                messagebox.showwarning(
                    "Existing Copy Session",
                    "This CSV already has a resumable copy session. Use Run CSV Bulk Copy to resume it, or restart the session first.",
                    parent=self.root,
                )
                return

        if checkpoint:
            resume_action = self._prompt_simplified_bulk_resume_action(checkpoint)
            if resume_action == "cancel":
                self._append_log("Simplified bulk dry run cancelled before session resume.")
                return
            if resume_action == "restart":
                self._delete_simplified_bulk_checkpoint(csv_path)
                checkpoint = self._build_new_simplified_bulk_checkpoint(csv_path, len(copy_items))
                self._append_log("Restarting Simplified Bulk Copy dry run from the beginning.")
            else:
                self._append_log(
                    f"Resuming Simplified Bulk Copy dry run from {self._checkpoint_phase_label(str(checkpoint.get('phase', 'dry_run'))).lower()}."
                )
        else:
            checkpoint = self._build_new_simplified_bulk_checkpoint(csv_path, len(copy_items))

        self._set_pause_state(False, active=False)
        self._set_running(True)
        self._append_log(f"Starting optional simplified bulk dry run for {len(copy_items)} row(s).")
        threading.Thread(
            target=self._simplified_bulk_preflight_and_copy_worker,
            args=(copy_items, checkpoint),
            daemon=True,
        ).start()

    def _on_inventory_clicked(self) -> None:
        inventory_uri = self.inventory_path_var.get().strip()
        try:
            _bucket, _prefix, normalized_uri = self._parse_s3_inventory_uri(inventory_uri)
        except ValueError as error:
            messagebox.showerror("Inventory Validation", str(error), parent=self.root)
            self._append_log(f"Inventory validation failed: {error}")
            return

        confirm_message = (
            "Inventory export will scan this S3 location and write a CSV report.\n\n"
            f"Location: {normalized_uri}\n\n"
            "This is read-only and will not copy, rename, or delete anything.\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Inventory Export", confirm_message, parent=self.root):
            self._append_log("Inventory export cancelled before execution.")
            return

        self._set_running(True)
        threading.Thread(
            target=self._inventory_worker,
            args=(normalized_uri,),
            daemon=True,
        ).start()

    def _start_bulk_copy(self, rows: list[dict[str, str]]) -> bool:
        if self._running:
            messagebox.showerror("Bulk Copy", "A copy is already running. Wait for it to finish first.", parent=self.root)
            return False

        all_errors: list[str] = []
        all_copy_items: list[tuple[str, ResolvedS3Paths]] = []
        resolved_title_count = 0

        for index, row in enumerate(rows, start=1):
            row_label = str(row.get("title", "")).strip() or f"Title {index}"

            user_input = sanitize_user_input(
                str(row.get("current_file_name", "")),
                str(row.get("desired_move_folder", "")),
                str(row.get("desired_name", "")),
                str(row.get("current_caption_name", "")),
                str(row.get("desired_caption_name", "")),
            )

            if not any(
                [
                    user_input.current_file_name,
                    user_input.desired_move_folder,
                    user_input.desired_name,
                    user_input.current_caption_name,
                    user_input.desired_caption_name,
                ]
            ):
                continue

            resolved_title_count += 1
            row_errors, row_copy_items = self._prepare_copy_items(user_input, row_label)
            if row_errors:
                all_errors.extend([f"{row_label}: {error}" for error in row_errors])
                continue

            all_copy_items.extend(row_copy_items)

        if resolved_title_count == 0:
            messagebox.showerror("Bulk Copy", "No populated rows found. Fill at least one row first.", parent=self.root)
            return False

        if all_errors:
            messagebox.showerror("Bulk Copy Validation", "\n".join(all_errors), parent=self.root)
            self._append_log(f"Bulk copy validation failed: {' | '.join(all_errors)}")
            return False

        object_count = len(all_copy_items)
        title_count = sum(1 for item_label, _ in all_copy_items if item_label.endswith("Primary file"))
        confirm_message = (
            f"Bulk copy is ready.\n\nTitles: {title_count}\nObjects to copy: {object_count}\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Bulk Copy", confirm_message, parent=self.root):
            self._append_log("Bulk copy cancelled before execution.")
            return False

        self._set_running(True)
        self._append_log(f"Starting bulk copy for {title_count} title(s), {object_count} object(s).")
        threading.Thread(target=self._copy_worker, args=(all_copy_items,), daemon=True).start()
        return True

    def _start_bulk_direct_upload(self, rows: list[dict[str, str]]) -> bool:
        if self._running:
            messagebox.showerror(
                "Bulk Direct Upload",
                "An operation is already running. Wait for it to finish first.",
                parent=self.root,
            )
            return False

        all_errors: list[str] = []
        all_upload_items: list[DirectUploadItem] = []
        resolved_title_count = 0

        for index, row in enumerate(rows, start=1):
            row_label = str(row.get("title", "")).strip() or f"Title {index}"
            local_file_path = str(row.get("local_file_path", ""))
            desired_move_folder = str(row.get("desired_move_folder", ""))
            desired_name = str(row.get("desired_name", ""))

            if not any(
                [
                    local_file_path.strip(),
                    desired_move_folder.strip(),
                    desired_name.strip(),
                ]
            ):
                continue

            resolved_title_count += 1
            row_errors, row_items = self._prepare_direct_upload_items(
                row_label,
                local_file_path,
                "",
                desired_move_folder,
                desired_name,
                "",
            )
            if row_errors:
                all_errors.extend([f"{row_label}: {error}" for error in row_errors])
                continue
            all_upload_items.extend(row_items)

        if resolved_title_count == 0:
            messagebox.showerror(
                "Bulk Direct Upload",
                "No populated rows found. Fill at least one row first.",
                parent=self.root,
            )
            return False

        if all_errors:
            messagebox.showerror("Bulk Direct Upload Validation", "\n".join(all_errors), parent=self.root)
            self._append_log(f"Bulk direct upload validation failed: {' | '.join(all_errors)}")
            return False

        object_count = len(all_upload_items)
        title_count = sum(1 for item in all_upload_items if item.label.endswith("Primary file"))
        confirm_message = (
            f"Bulk direct upload is ready.\n\nTitles: {title_count}\nObjects to upload: {object_count}\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Bulk Direct Upload", confirm_message, parent=self.root):
            self._append_log("Bulk direct upload cancelled before execution.")
            return False

        self._set_running(True)
        self._append_log(f"Starting bulk direct upload for {title_count} title(s), {object_count} object(s).")
        threading.Thread(target=self._upload_worker, args=(all_upload_items,), daemon=True).start()
        return True

    def _refresh_preview(self) -> None:
        is_direct_upload_mode = self._is_direct_upload_mode()
        is_rename_mode = self._is_rename_mode()
        is_simplified_bulk_mode = self._is_simplified_bulk_mode()
        is_inventory_mode = self._is_inventory_mode()
        is_folder_copy_mode = self._is_folder_copy_mode()
        is_bulk_folder_copy_mode = self._is_bulk_folder_copy_mode()
        self._update_pause_button_state()
        if is_rename_mode:
            self.copy_button.configure(text="Rename")
        elif is_simplified_bulk_mode:
            self.copy_button.configure(text="Run CSV Bulk Copy")
        elif is_inventory_mode:
            self.copy_button.configure(text="Export Inventory")
        elif is_folder_copy_mode:
            self.copy_button.configure(text="Run Folder Copy")
        elif is_bulk_folder_copy_mode:
            self.copy_button.configure(text="Run Bulk Folder Copy")
        else:
            self.copy_button.configure(text="Upload" if is_direct_upload_mode else "Copy")

        bulk_label = "Bulk Upload..." if is_direct_upload_mode else "Bulk Copy..."
        self.bulk_copy_button.configure(text=bulk_label)
        self.settings_menu.entryconfigure(self.bulk_menu_index, label=bulk_label)

        if is_rename_mode or is_simplified_bulk_mode or is_inventory_mode or is_folder_copy_mode or is_bulk_folder_copy_mode:
            self.desired_block.grid_remove()
            self.bulk_copy_button.configure(state="disabled")
            self.settings_menu.entryconfigure(self.bulk_menu_index, state="disabled")
            self.preview_source_caption_label.grid_remove()
            self.preview_source_caption_entry.grid_remove()
            self.preview_dest_caption_label.grid_remove()
            self.preview_dest_caption_entry.grid_remove()
        else:
            self.desired_block.grid()
            self.desired_caption_entry.configure(state="normal")
            self.desired_caption_label.configure(text="Desired Caption Name (optional)")
            if not self._running:
                self.bulk_copy_button.configure(state="normal")
            self.settings_menu.entryconfigure(self.bulk_menu_index, state="normal")
            self.preview_source_caption_label.grid(row=3, column=0, sticky="w", padx=(10, 8), pady=(0, 4))
            self.preview_source_caption_entry.grid(row=3, column=1, sticky="ew", pady=(0, 4))
            self.preview_dest_caption_label.grid(row=4, column=0, sticky="w", padx=(10, 8), pady=(0, 8))
            self.preview_dest_caption_entry.grid(row=4, column=1, sticky="ew", pady=(0, 8))

        desired_move_folder = sanitize_folder_path(self.desired_move_folder_var.get())
        desired_name = sanitize_filename(self.desired_name_var.get())
        desired_caption_name = sanitize_filename(self.desired_caption_name_var.get())

        dest_bucket = self.config.dest_bucket.strip()
        primary_dest_key = join_key_parts(self.config.dest_prefix, desired_move_folder, desired_name)
        caption_dest_key = join_key_parts(self.config.dest_prefix, desired_move_folder, desired_caption_name)

        if is_rename_mode:
            current_destination_path = sanitize_folder_path(self.rename_current_path_var.get())
            current_destination_name = sanitize_filename(self.rename_current_name_var.get())
            desired_rename_name = sanitize_filename(self.rename_desired_name_var.get())
            source_key = join_key_parts(self.config.dest_prefix, current_destination_path, current_destination_name)
            dest_key = join_key_parts(self.config.dest_prefix, current_destination_path, desired_rename_name)
            if dest_bucket and source_key:
                self.source_preview_var.set(f"s3://{dest_bucket}/{source_key}")
            else:
                self.source_preview_var.set("")
            self.dest_preview_var.set(f"s3://{dest_bucket}/{dest_key}" if dest_bucket and dest_key else "")
            self.source_caption_preview_var.set("")
            self.dest_caption_preview_var.set("")
            return

        if is_simplified_bulk_mode:
            self._update_simplified_bulk_summary()
            errors, copy_items, preview = self._load_simplified_bulk_copy_items(self.simplified_bulk_csv_path_var.get())
            if errors or not preview or not copy_items:
                self.source_preview_var.set("")
                self.dest_preview_var.set("")
            else:
                self.source_preview_var.set(preview.first_source_uri)
                self.dest_preview_var.set(preview.first_destination_uri)
            self.source_caption_preview_var.set("")
            self.dest_caption_preview_var.set("")
            return

        if is_inventory_mode:
            inventory_uri = self.inventory_path_var.get().strip()
            self.source_preview_var.set(inventory_uri)
            self.dest_preview_var.set("")
            self.source_caption_preview_var.set("")
            self.dest_caption_preview_var.set("")

            if not inventory_uri:
                self.inventory_summary_var.set(
                    "Enter an S3 bucket or prefix URI to export a CSV inventory of everything under that location."
                )
            else:
                try:
                    _bucket, prefix, normalized_uri = self._parse_s3_inventory_uri(inventory_uri)
                    scope_text = "bucket root" if not prefix else "prefix"
                    self.inventory_summary_var.set(f"Ready to export an inventory for {scope_text}: {normalized_uri}")
                    self.source_preview_var.set(normalized_uri)
                except ValueError as error:
                    self.inventory_summary_var.set(str(error))
            return

        if is_folder_copy_mode:
            source_folder_uri = self.folder_copy_source_uri_var.get().strip()
            dest_folder_uri = self.folder_copy_dest_uri_var.get().strip()
            self.source_preview_var.set(source_folder_uri)
            self.dest_preview_var.set(dest_folder_uri)
            self.source_caption_preview_var.set("")
            self.dest_caption_preview_var.set("")

            if not source_folder_uri and not dest_folder_uri:
                self.folder_copy_summary_var.set(
                    "Provide source and destination S3 folder URIs. Relative paths under the source folder will be preserved."
                )
            else:
                self.folder_copy_summary_var.set(
                    "Ready to scan the source folder and copy all contained objects while preserving relative paths."
                )
            return

        if is_bulk_folder_copy_mode:
            self._update_bulk_folder_summary()
            errors, folder_jobs, preview = self._load_bulk_folder_copy_jobs(self.bulk_folder_csv_path_var.get())
            if errors or not folder_jobs or preview is None:
                self.source_preview_var.set("")
                self.dest_preview_var.set("")
            else:
                self.source_preview_var.set(preview.first_source_uri)
                self.dest_preview_var.set(preview.first_destination_uri)
            self.source_caption_preview_var.set("")
            self.dest_caption_preview_var.set("")
            return

        if self._is_direct_upload_mode():
            self.source_preview_var.set(self.local_file_path_var.get().strip())
            self.dest_preview_var.set(f"s3://{dest_bucket}/{primary_dest_key}" if dest_bucket and primary_dest_key else "")
            self.source_caption_preview_var.set(self.local_caption_path_var.get().strip())
            if self.local_caption_path_var.get().strip() and desired_caption_name and dest_bucket and caption_dest_key:
                self.dest_caption_preview_var.set(f"s3://{dest_bucket}/{caption_dest_key}")
            else:
                self.dest_caption_preview_var.set("")
            return

        user_input = sanitize_user_input(
            self.current_file_name_var.get(),
            self.desired_move_folder_var.get(),
            self.desired_name_var.get(),
            self.current_caption_name_var.get(),
            self.desired_caption_name_var.get(),
        )
        paths = build_paths(self.config, user_input)
        caption_paths = self._build_caption_paths(user_input)

        self.source_preview_var.set(paths.source_uri)
        self.dest_preview_var.set(paths.dest_uri)
        if caption_paths:
            self.source_caption_preview_var.set(caption_paths.source_uri)
            self.dest_caption_preview_var.set(caption_paths.dest_uri)
        else:
            self.source_caption_preview_var.set("")
            self.dest_caption_preview_var.set("")

    def _set_running(self, running: bool) -> None:
        self._running = running
        if running:
            self._update_pause_button_state()
            self.copy_button.configure(state="disabled")
            self.bulk_copy_button.configure(state="disabled")
            self.settings_menu.entryconfigure(self.bulk_menu_index, state="disabled")
            if self.simplified_bulk_dry_run_button is not None:
                self.simplified_bulk_dry_run_button.configure(state="disabled")
        else:
            self.copy_button.configure(state="normal")
            if (
                self._is_rename_mode()
                or self._is_simplified_bulk_mode()
                or self._is_inventory_mode()
                or self._is_folder_copy_mode()
                or self._is_bulk_folder_copy_mode()
            ):
                self.bulk_copy_button.configure(state="disabled")
                self.settings_menu.entryconfigure(self.bulk_menu_index, state="disabled")
            else:
                self.bulk_copy_button.configure(state="normal")
                self.settings_menu.entryconfigure(self.bulk_menu_index, state="normal")
            self._set_pause_state(False, active=False)
            if self.simplified_bulk_dry_run_button is not None:
                self.simplified_bulk_dry_run_button.configure(state="normal")

    def _append_log(self, message: str) -> None:
        if self._closing:
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_area.configure(state="normal")
        self.log_area.insert("end", f"[{timestamp}] {message}\n")
        self.log_area.see("end")
        self.log_area.configure(state="disabled")

    def _play_completion_notification(self) -> None:
        if IS_WINDOWS:
            try:
                winsound.MessageBeep(winsound.MB_ICONASTERISK)
                return
            except Exception:  # pylint: disable=broad-except
                pass

        try:
            if IS_MACOS:
                subprocess.Popen(VOICE_COMMAND, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                return
        except Exception:  # pylint: disable=broad-except
            pass

        try:
            if IS_MACOS:
                subprocess.Popen(DING_COMMAND, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            elif not IS_WINDOWS:
                print("\a", end="", flush=True)
        except Exception:  # pylint: disable=broad-except
            pass

    def _enqueue_ui(self, callback, *args, **kwargs) -> None:
        if self._closing:
            return
        self._ui_queue.put((callback, args, kwargs))

    def _process_ui_queue(self) -> None:
        if self._closing:
            return
        while True:
            try:
                callback, args, kwargs = self._ui_queue.get_nowait()
            except queue.Empty:
                break
            callback(*args, **kwargs)

        self.root.after(100, self._process_ui_queue)

    def _collect_gc_on_ui_thread(self) -> None:
        if self._closing:
            return
        gc.collect()
        self.root.after(10_000, self._collect_gc_on_ui_thread)

    def _on_close(self) -> None:
        self._closing = True
        gc.enable()
        self.root.destroy()

    def _call_on_ui_thread(self, callback, *args, **kwargs):
        if threading.current_thread() is threading.main_thread():
            return callback(*args, **kwargs)

        done = threading.Event()
        result = {}

        def wrapper() -> None:
            try:
                result["value"] = callback(*args, **kwargs)
            except Exception as error:  # pylint: disable=broad-except
                result["error"] = error
            finally:
                done.set()

        self._enqueue_ui(wrapper)
        done.wait()

        if "error" in result:
            raise result["error"]
        return result.get("value")

    def _present_modal_dialog(self, dialog: tk.Toplevel) -> None:
        self.root.update_idletasks()
        dialog.update_idletasks()

        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_width = max(self.root.winfo_width(), 1)
        root_height = max(self.root.winfo_height(), 1)
        dialog_width = max(dialog.winfo_reqwidth(), dialog.winfo_width(), 1)
        dialog_height = max(dialog.winfo_reqheight(), dialog.winfo_height(), 1)

        centered_x = root_x + max((root_width - dialog_width) // 2, 0)
        centered_y = root_y + max((root_height - dialog_height) // 2, 0)
        dialog.geometry(f"+{centered_x}+{centered_y}")
        dialog.deiconify()
        dialog.lift()
        dialog.focus_force()

    def _build_caption_paths(self, user_input: UserInput) -> ResolvedS3Paths | None:
        if not user_input.current_caption_name or not user_input.desired_caption_name:
            return None

        caption_input = replace(
            user_input,
            current_file_name=user_input.current_caption_name,
            desired_name=user_input.desired_caption_name,
        )
        return build_paths(self.config, caption_input)

    def _copy_one_object(
        self,
        s3_client,
        item_label: str,
        paths: ResolvedS3Paths,
        overwrite_mode: str = "prompt",
    ) -> None:
        source_ref = S3ObjectRef(bucket=paths.source_bucket, key=paths.source_key)
        dest_ref = S3ObjectRef(bucket=paths.dest_bucket, key=paths.dest_key)

        self._enqueue_ui(self._append_log, f"Starting {item_label} copy: {paths.source_uri} -> {paths.dest_uri}")

        if not object_exists(s3_client, source_ref):
            raise UserVisibleError(f"{item_label} source file not found. Verify names and try again.")

        destination_exists = object_exists(s3_client, dest_ref)
        allow_overwrite = False
        if destination_exists:
            if overwrite_mode == "overwrite_all":
                self._enqueue_ui(
                    self._append_log,
                    f"{item_label} destination already exists. Proceeding because Overwrite All was approved.",
                )
                allow_overwrite = True
            elif overwrite_mode == "collect_overwrites":
                raise DeferredOverwriteError(
                    f"{item_label} destination already exists. Deferred for end-of-run overwrite review."
                )
            elif overwrite_mode == "deny_existing":
                raise UserVisibleError(
                    f"{item_label} destination already exists. Dry run completed without overwrite approval, so no copy was performed."
                )
            else:
                self._enqueue_ui(
                    self._append_log,
                    f"{item_label} destination already exists. Waiting for overwrite confirmation.",
                )
                should_proceed = self._call_on_ui_thread(
                    messagebox.askyesno,
                    "Destination Exists",
                    (
                        f"{item_label} destination object already exists.\n\n"
                        f"Destination: {paths.dest_uri}\n\n"
                        "Copying now will overwrite that object. Continue?"
                    ),
                    parent=self.root,
                )
                if not should_proceed:
                    raise UserVisibleError(
                        f"{item_label} copy cancelled. Destination object already exists and overwrite was not approved."
                    )
                allow_overwrite = True

        try:
            copy_object(
                s3_client,
                source_ref,
                dest_ref,
                allow_overwrite=allow_overwrite,
                progress_callback=lambda msg: self._enqueue_ui(self._append_log, f"{item_label}: {msg}"),
            )
        except DestinationExistsError as error:
            raise UserVisibleError(
                f"{item_label} destination was created while copy was starting. No overwrite was performed. "
                "Review destination and run copy again only if you want to replace it."
            ) from error

    @staticmethod
    def _overwrite_mode_for_simplified_bulk_row(checkpoint: dict, row_status: str) -> str:
        configured_mode = str(checkpoint.get("overwrite_mode", "deny_existing") or "deny_existing")
        if configured_mode == "collect_overwrites":
            return "collect_overwrites"

        selection = str(checkpoint.get("copy_selection", "") or "").strip()
        if selection == "all_rows":
            return "overwrite_all"
        if selection == "include_overwrites":
            return "overwrite_all" if row_status == "overwrite_warning" else "deny_existing"
        if selection == "include_internal_conflicts":
            return "overwrite_all" if row_status == "internal_conflict" else "deny_existing"
        return configured_mode

    def _upload_one_object(self, s3_client, item: DirectUploadItem) -> None:
        self._enqueue_ui(self._append_log, f"Starting {item.label} upload: {item.local_path} -> {item.destination_uri}")

        if not os.path.isfile(item.local_path):
            raise UserVisibleError(f"{item.label} local file not found: {item.local_path}")

        destination_exists = object_exists(s3_client, item.destination_ref)
        if destination_exists:
            self._enqueue_ui(
                self._append_log,
                f"{item.label} destination already exists. Waiting for overwrite confirmation.",
            )
            should_proceed = self._call_on_ui_thread(
                messagebox.askyesno,
                "Destination Exists",
                (
                    f"{item.label} destination object already exists.\n\n"
                    f"Destination: {item.destination_uri}\n\n"
                    "Uploading now will overwrite that object. Continue?"
                ),
                parent=self.root,
            )
            if not should_proceed:
                raise UserVisibleError(
                    f"{item.label} upload cancelled. Destination object already exists and overwrite was not approved."
                )

        upload_local_file(
            s3_client,
            item.local_path,
            item.destination_ref,
            progress_callback=lambda msg: self._enqueue_ui(self._append_log, f"{item.label}: {msg}"),
        )

    def _rename_one_object(
        self,
        s3_client,
        source_ref: S3ObjectRef,
        destination_ref: S3ObjectRef,
        source_uri: str,
        dest_uri: str,
    ) -> None:
        self._enqueue_ui(self._append_log, f"Starting rename: {source_uri} -> {dest_uri}")

        if not object_exists(s3_client, source_ref):
            raise UserVisibleError("Current destination object was not found. Verify name/folder and try again.")

        destination_exists = object_exists(s3_client, destination_ref)
        allow_overwrite = False
        if destination_exists:
            self._enqueue_ui(
                self._append_log,
                "Rename destination already exists. Waiting for overwrite confirmation.",
            )
            should_proceed = self._call_on_ui_thread(
                messagebox.askyesno,
                "Destination Exists",
                (
                    "Rename destination object already exists.\n\n"
                    f"Destination: {dest_uri}\n\n"
                    "Renaming now will overwrite that object. Continue?"
                ),
                parent=self.root,
            )
            if not should_proceed:
                raise UserVisibleError("Rename cancelled. Destination object already exists and overwrite was not approved.")
            allow_overwrite = True

        try:
            copy_object(
                s3_client,
                source_ref,
                destination_ref,
                allow_overwrite=allow_overwrite,
                progress_callback=lambda msg: self._enqueue_ui(self._append_log, f"Rename: {msg}"),
            )
        except DestinationExistsError as error:
            raise UserVisibleError(
                "Rename destination was created while rename was starting. No overwrite was performed."
            ) from error

        try:
            delete_object(s3_client, source_ref)
        except UserVisibleError as error:
            raise UserVisibleError(
                "Rename copied the destination, but failed to remove the original source object. "
                "Both objects may now exist; review and clean up manually."
            ) from error

    def on_copy_clicked(self) -> None:
        if self._running:
            return

        if self._is_rename_mode():
            self._on_rename_clicked()
            return

        if self._is_simplified_bulk_mode():
            self._on_simplified_bulk_copy_clicked()
            return

        if self._is_inventory_mode():
            self._on_inventory_clicked()
            return

        if self._is_folder_copy_mode():
            self._on_folder_copy_clicked()
            return

        if self._is_bulk_folder_copy_mode():
            self._on_bulk_folder_copy_clicked()
            return

        if self._is_direct_upload_mode():
            self._on_direct_upload_clicked()
            return

        user_input = sanitize_user_input(
            self.current_file_name_var.get(),
            self.desired_move_folder_var.get(),
            self.desired_name_var.get(),
            self.current_caption_name_var.get(),
            self.desired_caption_name_var.get(),
        )

        validation_errors, copy_items = self._prepare_copy_items(user_input, "Single copy")

        if validation_errors:
            messagebox.showerror("Validation", "\n".join(validation_errors), parent=self.root)
            self._append_log(f"Validation failed: {' | '.join(validation_errors)}")
            return

        path_lines = []
        for item_label, item_paths in copy_items:
            path_lines.append(f"{item_label} Source: {item_paths.source_uri}")
            path_lines.append(f"{item_label} Destination: {item_paths.dest_uri}")

        confirm_message = (
            "Copy will run with these resolved paths:\n\n"
            + "\n".join(path_lines)
            + "\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Copy", confirm_message, parent=self.root):
            self._append_log("Copy cancelled before execution.")
            return

        self._set_running(True)
        threading.Thread(target=self._copy_worker, args=(copy_items,), daemon=True).start()

    def _on_direct_upload_clicked(self) -> None:
        errors, upload_items = self._prepare_direct_upload_items(
            "Direct upload",
            self.local_file_path_var.get(),
            self.local_caption_path_var.get(),
            self.desired_move_folder_var.get(),
            self.desired_name_var.get(),
            self.desired_caption_name_var.get(),
        )
        if errors:
            messagebox.showerror("Validation", "\n".join(errors), parent=self.root)
            self._append_log(f"Validation failed: {' | '.join(errors)}")
            return

        path_lines = []
        for item in upload_items:
            path_lines.append(f"{item.label} Source: {item.local_path}")
            path_lines.append(f"{item.label} Destination: {item.destination_uri}")

        confirm_message = (
            "Direct upload will run with these resolved paths:\n\n"
            + "\n".join(path_lines)
            + "\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Direct Upload", confirm_message, parent=self.root):
            self._append_log("Direct upload cancelled before execution.")
            return

        self._set_running(True)
        threading.Thread(target=self._upload_worker, args=(upload_items,), daemon=True).start()

    def _on_rename_clicked(self) -> None:
        errors, rename_data = self._prepare_rename_item()
        if errors or not rename_data:
            messagebox.showerror("Validation", "\n".join(errors), parent=self.root)
            self._append_log(f"Validation failed: {' | '.join(errors)}")
            return

        source_ref, destination_ref, source_uri, dest_uri = rename_data
        confirm_message = (
            "Rename will run with these resolved destination paths:\n\n"
            f"Current: {source_uri}\n"
            f"Renamed: {dest_uri}\n\n"
            "This operation copies to the new key and then deletes the old key.\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Rename", confirm_message, parent=self.root):
            self._append_log("Rename cancelled before execution.")
            return

        self._set_running(True)
        threading.Thread(
            target=self._rename_worker,
            args=(source_ref, destination_ref, source_uri, dest_uri),
            daemon=True,
        ).start()

    def _on_folder_copy_clicked(self) -> None:
        source_folder_uri = self.folder_copy_source_uri_var.get().strip()
        destination_folder_uri = self.folder_copy_dest_uri_var.get().strip()

        errors: list[str] = []
        try:
            normalized_source_uri = self._parse_s3_folder_uri(source_folder_uri)[2]
        except ValueError as error:
            normalized_source_uri = ""
            errors.append(f"Source Folder S3 URI: {error}")

        try:
            normalized_destination_uri = self._parse_s3_folder_uri(destination_folder_uri)[2]
        except ValueError as error:
            normalized_destination_uri = ""
            errors.append(f"Destination Folder S3 URI: {error}")

        if (
            normalized_source_uri
            and normalized_destination_uri
            and normalized_source_uri == normalized_destination_uri
        ):
            errors.append("Source folder and destination folder resolve to the same S3 location.")

        if errors:
            messagebox.showerror("Validation", "\n".join(errors), parent=self.root)
            self._append_log(f"Validation failed: {' | '.join(errors)}")
            return

        confirm_message = (
            "Folder copy will scan the source folder and copy every object it finds.\n\n"
            f"Source Folder: {normalized_source_uri}\n"
            f"Destination Folder: {normalized_destination_uri}\n\n"
            "Relative paths under the source folder will be preserved.\n"
            "Destination overwrite candidates will be reviewed once at the end.\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Folder Copy", confirm_message, parent=self.root):
            self._append_log("Folder copy cancelled before execution.")
            return

        self._set_running(True)
        threading.Thread(
            target=self._folder_copy_worker,
            args=(normalized_source_uri, normalized_destination_uri),
            daemon=True,
        ).start()

    def _on_bulk_folder_copy_clicked(self) -> None:
        csv_path = self.bulk_folder_csv_path_var.get().strip()
        errors, folder_jobs, preview = self._load_bulk_folder_copy_jobs(csv_path)
        if errors:
            messagebox.showerror("Validation", "\n".join(errors), parent=self.root)
            self._append_log(f"Validation failed: {' | '.join(errors)}")
            return

        if not folder_jobs or preview is None:
            messagebox.showerror(
                "Validation",
                "No valid folder pairs were found in the CSV file.",
                parent=self.root,
            )
            self._append_log("Validation failed: no valid folder pairs were found in the CSV file.")
            return

        confirm_message = (
            "Bulk folder copy will scan each source folder and copy every object it finds.\n\n"
            f"CSV File: {Path(csv_path).name}\n"
            f"Folder pairs: {preview.folder_pair_count}\n"
            f"First Source Folder: {preview.first_source_uri}\n"
            f"First Destination Folder: {preview.first_destination_uri}\n\n"
            "Relative paths under each source folder will be preserved.\n"
            "Destination overwrite candidates will be reviewed once at the end.\n\n"
            "Continue?"
        )
        if not messagebox.askokcancel("Confirm Bulk Folder Copy", confirm_message, parent=self.root):
            self._append_log("Bulk folder copy cancelled before execution.")
            return

        self._set_running(True)
        threading.Thread(
            target=self._bulk_folder_copy_worker,
            args=(folder_jobs,),
            daemon=True,
        ).start()

    def _copy_worker(self, copy_items: list[tuple[str, ResolvedS3Paths]]) -> None:

        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            for item_label, item_paths in copy_items:
                self._copy_one_object(s3_client, item_label, item_paths)

            self._enqueue_ui(self._append_log, "Copy succeeded. Source object(s) were not deleted.")
            self._play_completion_notification()
            success_lines = [
                f"Objects copied: {len(copy_items)}",
            ]
            if len(copy_items) <= 6:
                for item_label, item_paths in copy_items:
                    success_lines.append(f"{item_label} Source: {item_paths.source_uri}")
                    success_lines.append(f"{item_label} Destination: {item_paths.dest_uri}")
            else:
                success_lines.append("See Status / Output for per-item details.")
            self._enqueue_ui(
                messagebox.showinfo,
                "Success",
                (
                    "Copy completed successfully.\n\n"
                    + "\n".join(success_lines)
                ),
                parent=self.root,
            )

        except UserVisibleError as error:
            self._enqueue_ui(self._append_log, f"Copy failed: {error}")
            self._enqueue_ui(messagebox.showerror, "Copy Failed", str(error), parent=self.root)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Copy Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)

    def _simplified_bulk_preflight_and_copy_worker(
        self,
        copy_items: list[tuple[str, ResolvedS3Paths]],
        checkpoint: dict,
    ) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            aws_cli_path = self._find_aws_cli_path()
            rows_checked = len(copy_items)

            if aws_cli_path:
                self._enqueue_ui(self._append_log, f"Using AWS CLI dry run at {aws_cli_path}")
            else:
                self._enqueue_ui(
                    self._append_log,
                    "AWS CLI dry run was not available on this Mac; internal validation will be used.",
                )

            if checkpoint.get("phase") != "awaiting_confirmation":
                checkpoint["phase"] = "dry_run"
                start_index = int(checkpoint.get("next_index", 0))
                duplicate_destination_indices = self._find_duplicate_destination_indices(copy_items)
                folder_existence_cache: dict[tuple[str, str], bool] = {}
                if duplicate_destination_indices and start_index == 0:
                    self._enqueue_ui(
                        self._append_log,
                        (
                            f"Found {len(duplicate_destination_indices)} row(s) with duplicate destination targets in the CSV. "
                            "Those rows will be marked as internal conflicts in the dry-run report."
                        ),
                    )
                for index in range(start_index, len(copy_items)):
                    self._wait_if_simplified_bulk_paused(checkpoint, "Dry run")
                    item_label, item_paths = copy_items[index]
                    source_ref, dest_ref = self._build_s3_ref_from_paths(item_paths)
                    destination_folder_uri = ""
                    destination_folder_status = ""
                    destination_folder_message = ""

                    try:
                        (
                            destination_folder_uri,
                            destination_folder_status,
                            destination_folder_message,
                        ) = self._resolve_destination_folder_status(
                            s3_client,
                            item_paths,
                            folder_existence_cache,
                        )
                        if not object_exists(s3_client, source_ref):
                            raise UserVisibleError("Source file not found.")

                        destination_exists = object_exists(s3_client, dest_ref)
                        cli_message = "Internal validation only."
                        if aws_cli_path:
                            cli_ok, cli_output = self._run_aws_cli_dry_run(
                                aws_cli_path,
                                credentials,
                                item_paths.source_uri,
                                item_paths.dest_uri,
                            )
                            if not cli_ok:
                                raise UserVisibleError(cli_output)
                            cli_message = cli_output

                        if index in duplicate_destination_indices:
                            row = SimplifiedBulkCopyReportRow(
                                row_label=item_label,
                                source_uri=item_paths.source_uri,
                                destination_uri=item_paths.dest_uri,
                                destination_folder_uri=destination_folder_uri,
                                destination_folder_status=destination_folder_status,
                                status="internal_conflict",
                                message=(
                                    "Another row in this CSV targets the same destination object. "
                                    f"{destination_folder_message} {cli_message}"
                                ).strip(),
                            )
                            self._enqueue_ui(self._append_log, f"{item_label}: dry run internal conflict.")
                        elif destination_exists:
                            row = SimplifiedBulkCopyReportRow(
                                row_label=item_label,
                                source_uri=item_paths.source_uri,
                                destination_uri=item_paths.dest_uri,
                                destination_folder_uri=destination_folder_uri,
                                destination_folder_status=destination_folder_status,
                                status="overwrite_warning",
                                message=f"Destination already exists. {destination_folder_message} {cli_message}",
                            )
                            self._enqueue_ui(
                                self._append_log,
                                f"{item_label}: dry run warning. Destination already exists.",
                            )
                        else:
                            row = SimplifiedBulkCopyReportRow(
                                row_label=item_label,
                                source_uri=item_paths.source_uri,
                                destination_uri=item_paths.dest_uri,
                                destination_folder_uri=destination_folder_uri,
                                destination_folder_status=destination_folder_status,
                                status="ready",
                                message=f"{destination_folder_message} {cli_message}",
                            )
                            self._enqueue_ui(self._append_log, f"{item_label}: dry run passed.")
                    except (UserVisibleError, RuntimeError) as error:
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            destination_folder_uri=destination_folder_uri,
                            destination_folder_status=destination_folder_status,
                            status="failed",
                            message=f"{error} {destination_folder_message}".strip(),
                        )
                        self._enqueue_ui(self._append_log, f"{item_label}: dry run failed: {error}")
                    except Exception as error:  # pylint: disable=broad-except
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            destination_folder_uri=destination_folder_uri,
                            destination_folder_status=destination_folder_status,
                            status="failed",
                            message=f"Unexpected error: {error} {destination_folder_message}".strip(),
                        )
                        self._enqueue_ui(self._append_log, f"{item_label}: dry run unexpected failure: {error}")

                    checkpoint.setdefault("dry_run_rows", []).append(asdict(row))
                    checkpoint["next_index"] = index + 1
                    self._save_simplified_bulk_checkpoint(checkpoint)
                    self._sync_simplified_bulk_report(checkpoint, "dry_run")

            checkpoint["phase"] = "awaiting_confirmation"
            checkpoint["paused"] = False
            self._save_simplified_bulk_checkpoint(checkpoint)
            dry_run_report_path = self._sync_simplified_bulk_report(checkpoint, "dry_run")
            ready_count, overwrite_count, internal_conflict_count, error_count = self._summarize_simplified_bulk_rows(
                checkpoint.get("dry_run_rows", [])
            )
            self._enqueue_ui(
                self._append_log,
                f"Simplified bulk dry run finished. Report written to {dry_run_report_path}",
            )
            action = self._call_on_ui_thread(
                self._prompt_simplified_bulk_preflight_action,
                dry_run_report_path,
                rows_checked,
                ready_count,
                overwrite_count,
                internal_conflict_count,
                error_count,
            )

            if action == "cancel":
                self._enqueue_ui(self._append_log, "Simplified bulk copy cancelled after dry run review.")
                return

            if action == "include_overwrites":
                overwrite_mode = "overwrite_all"
                copy_selection = "include_overwrites"
            elif action == "include_internal_conflicts":
                overwrite_mode = "deny_existing"
                copy_selection = "include_internal_conflicts"
            elif action == "all_rows":
                overwrite_mode = "overwrite_all"
                copy_selection = "all_rows"
            else:
                overwrite_mode = "deny_existing"
                copy_selection = "ready_only"
            checkpoint["phase"] = "copy"
            checkpoint["overwrite_mode"] = overwrite_mode
            checkpoint["copy_selection"] = copy_selection
            planned_items = self._build_simplified_bulk_copy_plan(copy_items, checkpoint)
            if not planned_items:
                self._enqueue_ui(
                    self._append_log,
                    "Dry run finished, but there were no eligible rows to copy after exclusions.",
                )
                self._enqueue_ui(
                    messagebox.showwarning,
                    "No Rows To Copy",
                    "Dry run finished, but there are no eligible rows to copy. Review the report and try again.",
                    parent=self.root,
                )
                return

            checkpoint["next_index"] = int(len(checkpoint.get("copy_rows", [])))
            checkpoint["paused"] = False
            self._save_simplified_bulk_checkpoint(checkpoint)
            self._enqueue_ui(
                self._append_log,
                f"Dry run approved. Starting actual simplified bulk copy for {len(planned_items)} selected row(s).",
            )
            self._simplified_bulk_copy_worker(planned_items, checkpoint, clear_running=False)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Simplified Bulk Copy Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)

    def _simplified_bulk_copy_worker(
        self,
        copy_items: list[tuple[int, str, ResolvedS3Paths, str]],
        checkpoint: dict,
        clear_running: bool = True,
    ) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            phase = str(checkpoint.get("phase", "copy"))
            overwrite_mode = str(checkpoint.get("overwrite_mode", "deny_existing") or "deny_existing")

            if phase == "copy_overwrite":
                pending_entries = self._find_pending_overwrite_entries(copy_items, checkpoint)
                start_index = int(checkpoint.get("overwrite_next_index", 0))
                for pending_offset in range(start_index, len(pending_entries)):
                    self._wait_if_simplified_bulk_paused(checkpoint, "Overwrite pass")
                    row_index, _original_index, item_label, item_paths, _row_status = pending_entries[pending_offset]
                    try:
                        self._copy_one_object(s3_client, item_label, item_paths, overwrite_mode="overwrite_all")
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="success",
                            message="Copied successfully after overwrite approval.",
                        )
                    except (UserVisibleError, RuntimeError) as error:
                        self._enqueue_ui(self._append_log, f"{item_label} overwrite failed: {error}")
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="failed",
                            message=str(error),
                        )
                    except Exception as error:  # pylint: disable=broad-except
                        self._enqueue_ui(self._append_log, f"{item_label} overwrite unexpected failure: {error}")
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="failed",
                            message=f"Unexpected error: {error}",
                        )

                    checkpoint["copy_rows"][row_index] = asdict(row)
                    checkpoint["overwrite_next_index"] = pending_offset + 1
                    checkpoint["paused"] = False
                    self._save_simplified_bulk_checkpoint(checkpoint)
                    report_path = self._sync_simplified_bulk_report(checkpoint, "copy")
            else:
                checkpoint["phase"] = "copy"
                start_index = int(checkpoint.get("next_index", 0))

                for index in range(start_index, len(copy_items)):
                    self._wait_if_simplified_bulk_paused(checkpoint, "Copy")
                    _original_index, item_label, item_paths, row_status = copy_items[index]
                    row_overwrite_mode = self._overwrite_mode_for_simplified_bulk_row(checkpoint, row_status)
                    try:
                        self._copy_one_object(s3_client, item_label, item_paths, overwrite_mode=row_overwrite_mode)
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="success",
                            message="Copied successfully.",
                        )
                    except DeferredOverwriteError as error:
                        self._enqueue_ui(self._append_log, f"{item_label} queued for overwrite review.")
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="overwrite_pending",
                            message=str(error),
                        )
                    except (UserVisibleError, RuntimeError) as error:
                        self._enqueue_ui(self._append_log, f"{item_label} failed: {error}")
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="failed",
                            message=str(error),
                        )
                    except Exception as error:  # pylint: disable=broad-except
                        self._enqueue_ui(self._append_log, f"{item_label} unexpected failure: {error}")
                        row = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="failed",
                            message=f"Unexpected error: {error}",
                        )

                    checkpoint.setdefault("copy_rows", []).append(asdict(row))
                    checkpoint["next_index"] = index + 1
                    checkpoint["paused"] = False
                    self._save_simplified_bulk_checkpoint(checkpoint)
                    report_path = self._sync_simplified_bulk_report(checkpoint, "copy")

                if overwrite_mode == "collect_overwrites":
                    report_rows = self._report_rows_from_checkpoint(checkpoint.get("copy_rows", []))
                    success_count = sum(1 for row in report_rows if row.status == "success")
                    pending_count = sum(1 for row in report_rows if row.status == "overwrite_pending")
                    failure_count = sum(1 for row in report_rows if row.status == "failed")
                    report_path = self._sync_simplified_bulk_report(checkpoint, "copy")

                    if pending_count:
                        checkpoint["phase"] = "copy_overwrite_review"
                        checkpoint["overwrite_next_index"] = 0
                        self._save_simplified_bulk_checkpoint(checkpoint)
                        action = self._call_on_ui_thread(
                            self._prompt_simplified_bulk_copy_overwrite_action,
                            report_path,
                            success_count,
                            pending_count,
                            failure_count,
                        )
                        if action == "cancel":
                            self._enqueue_ui(
                                self._append_log,
                                "Simplified bulk copy paused at final overwrite review. Resume later to decide.",
                            )
                            return
                        if action == "finish_without_overwrites":
                            for row in checkpoint.get("copy_rows", []):
                                if str(row.get("status", "")) == "overwrite_pending":
                                    row["status"] = "skipped_overwrite"
                                    row["message"] = "Destination exists. Skipped because overwrite was not approved."
                            self._save_simplified_bulk_checkpoint(checkpoint)
                            report_path = self._sync_simplified_bulk_report(checkpoint, "copy")
                        else:
                            checkpoint["phase"] = "copy_overwrite"
                            checkpoint["overwrite_next_index"] = 0
                            self._save_simplified_bulk_checkpoint(checkpoint)
                            self._enqueue_ui(
                                self._append_log,
                                f"Starting overwrite pass for {pending_count} queued row(s).",
                            )
                            self._simplified_bulk_copy_worker(copy_items, checkpoint, clear_running=False)
                            return

            report_rows = self._report_rows_from_checkpoint(checkpoint.get("copy_rows", []))
            report_path = self._sync_simplified_bulk_report(checkpoint, "copy")
            success_count = sum(1 for row in report_rows if row.status == "success")
            skipped_overwrite_count = sum(1 for row in report_rows if row.status == "skipped_overwrite")
            failure_count = sum(1 for row in report_rows if row.status == "failed")

            if failure_count == 0 and skipped_overwrite_count == 0:
                self._enqueue_ui(
                    self._append_log,
                    f"Simplified bulk copy finished. Report written to {report_path}",
                )
                self._play_completion_notification()
                self._enqueue_ui(
                    messagebox.showinfo,
                    "Simplified Bulk Copy Complete",
                    (
                        f"Rows copied successfully: {success_count}\n"
                        f"Rows failed: {failure_count}\n\n"
                        f"Report saved to:\n{report_path}"
                    ),
                    parent=self.root,
                )
            elif failure_count > 0 or skipped_overwrite_count > 0:
                self._enqueue_ui(
                    self._append_log,
                    f"Simplified bulk copy finished with exceptions. Report written to {report_path}",
                )
                dialog_title = (
                    "Simplified Bulk Copy Finished With Failures"
                    if failure_count > 0
                    else "Simplified Bulk Copy Finished Without Overwrites"
                )
                self._enqueue_ui(
                    messagebox.showwarning,
                    dialog_title,
                    (
                        f"Rows copied successfully: {success_count}\n"
                        f"Rows failed: {failure_count}\n"
                        f"Skipped overwrite rows: {skipped_overwrite_count}\n\n"
                        f"Report saved to:\n{report_path}\n\n"
                        "Review the CSV report for per-row results."
                    ),
                    parent=self.root,
                )
            self._delete_simplified_bulk_checkpoint(str(checkpoint["csv_path"]))
            self._enqueue_ui(self._update_simplified_bulk_summary)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Simplified Bulk Copy Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            if clear_running:
                self._enqueue_ui(self._set_running, False)

    def _prompt_folder_copy_overwrite_action(
        self,
        report_path: Path,
        success_count: int,
        pending_count: int,
        failure_count: int,
    ) -> str:
        result = {"value": "finish_without_overwrites"}
        dialog = tk.Toplevel(self.root)
        dialog.title("Folder Copy Overwrite Review")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        body = ttk.Frame(dialog, padding=16)
        body.pack(fill="both", expand=True)

        detail_lines = [
            f"Copied successfully: {success_count}",
            f"Would overwrite existing destination files: {pending_count}",
            f"Failed: {failure_count}",
            "",
            "Open the report to review the overwrite candidates.",
            "You can overwrite those files now or finish without overwriting them.",
        ]

        ttk.Label(body, text="Folder copy main pass finished.").pack(anchor="w")
        ttk.Label(body, text="\n".join(detail_lines), justify="left").pack(anchor="w", pady=(8, 12))
        ttk.Label(body, text=f"Report: {report_path}", justify="left", wraplength=640).pack(anchor="w")

        button_row = ttk.Frame(body)
        button_row.pack(anchor="e", pady=(14, 0))

        ttk.Button(
            button_row,
            text="Open Report",
            command=lambda: self._open_report_file(report_path),
        ).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(
            button_row,
            text="Overwrite These Files",
            command=lambda: (result.__setitem__("value", "overwrite_all_pending"), dialog.destroy()),
        ).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(
            button_row,
            text="Finish Without Overwrites",
            command=lambda: dialog.destroy(),
        ).grid(row=0, column=2)

        self._present_modal_dialog(dialog)
        dialog.wait_window()
        return result["value"]

    def _execute_folder_copy_items(
        self,
        s3_client,
        copy_items: list[tuple[str, ResolvedS3Paths]],
        report_kind: str,
        operation_label: str,
    ) -> None:
        report_rows: list[SimplifiedBulkCopyReportRow] = []
        report_path = self._simplified_bulk_report_path(report_kind)

        pending_entries: list[tuple[int, str, ResolvedS3Paths]] = []
        for row_index, (item_label, item_paths) in enumerate(copy_items):
            try:
                self._copy_one_object(s3_client, item_label, item_paths, overwrite_mode="collect_overwrites")
                row = SimplifiedBulkCopyReportRow(
                    row_label=item_label,
                    source_uri=item_paths.source_uri,
                    destination_uri=item_paths.dest_uri,
                    status="success",
                    message="Copied successfully.",
                )
            except DeferredOverwriteError as error:
                self._enqueue_ui(self._append_log, f"{item_label} queued for end-of-run overwrite review.")
                row = SimplifiedBulkCopyReportRow(
                    row_label=item_label,
                    source_uri=item_paths.source_uri,
                    destination_uri=item_paths.dest_uri,
                    status="overwrite_pending",
                    message=str(error),
                )
                pending_entries.append((row_index, item_label, item_paths))
            except (UserVisibleError, RuntimeError) as error:
                self._enqueue_ui(self._append_log, f"{item_label} failed: {error}")
                row = SimplifiedBulkCopyReportRow(
                    row_label=item_label,
                    source_uri=item_paths.source_uri,
                    destination_uri=item_paths.dest_uri,
                    status="failed",
                    message=str(error),
                )
            except Exception as error:  # pylint: disable=broad-except
                self._enqueue_ui(self._append_log, f"{item_label} unexpected failure: {error}")
                row = SimplifiedBulkCopyReportRow(
                    row_label=item_label,
                    source_uri=item_paths.source_uri,
                    destination_uri=item_paths.dest_uri,
                    status="failed",
                    message=f"Unexpected error: {error}",
                )

            report_rows.append(row)
            report_path = self._write_simplified_bulk_report(report_rows, report_kind, report_path=report_path)

        success_count = sum(1 for row in report_rows if row.status == "success")
        pending_count = sum(1 for row in report_rows if row.status == "overwrite_pending")
        failure_count = sum(1 for row in report_rows if row.status == "failed")

        if pending_count:
            action = self._call_on_ui_thread(
                self._prompt_folder_copy_overwrite_action,
                report_path,
                success_count,
                pending_count,
                failure_count,
            )

            if action == "overwrite_all_pending":
                for row_index, item_label, item_paths in pending_entries:
                    try:
                        self._copy_one_object(s3_client, item_label, item_paths, overwrite_mode="overwrite_all")
                        report_rows[row_index] = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="success",
                            message="Copied successfully after overwrite approval.",
                        )
                    except (UserVisibleError, RuntimeError) as error:
                        self._enqueue_ui(self._append_log, f"{item_label} overwrite failed: {error}")
                        report_rows[row_index] = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="failed",
                            message=str(error),
                        )
                    except Exception as error:  # pylint: disable=broad-except
                        self._enqueue_ui(self._append_log, f"{item_label} overwrite unexpected failure: {error}")
                        report_rows[row_index] = SimplifiedBulkCopyReportRow(
                            row_label=item_label,
                            source_uri=item_paths.source_uri,
                            destination_uri=item_paths.dest_uri,
                            status="failed",
                            message=f"Unexpected error: {error}",
                        )
                    report_path = self._write_simplified_bulk_report(
                        report_rows,
                        report_kind,
                        report_path=report_path,
                    )
            else:
                for row_index, item_label, item_paths in pending_entries:
                    report_rows[row_index] = SimplifiedBulkCopyReportRow(
                        row_label=item_label,
                        source_uri=item_paths.source_uri,
                        destination_uri=item_paths.dest_uri,
                        status="skipped_overwrite",
                        message="Destination exists. Skipped because overwrite was not approved.",
                    )
                report_path = self._write_simplified_bulk_report(
                    report_rows,
                    report_kind,
                    report_path=report_path,
                )

        success_count = sum(1 for row in report_rows if row.status == "success")
        skipped_overwrite_count = sum(1 for row in report_rows if row.status == "skipped_overwrite")
        failure_count = sum(1 for row in report_rows if row.status == "failed")

        if failure_count == 0 and skipped_overwrite_count == 0:
            self._enqueue_ui(
                self._append_log,
                f"{operation_label} finished. Report written to {report_path}",
            )
            self._play_completion_notification()
            self._enqueue_ui(
                messagebox.showinfo,
                f"{operation_label} Complete",
                (
                    f"Objects copied successfully: {success_count}\n"
                    f"Objects failed: {failure_count}\n\n"
                    f"Report saved to:\n{report_path}"
                ),
                parent=self.root,
            )
        else:
            self._enqueue_ui(
                self._append_log,
                f"{operation_label} finished with exceptions. Report written to {report_path}",
            )
            self._enqueue_ui(
                messagebox.showwarning,
                f"{operation_label} Finished With Exceptions",
                (
                    f"Objects copied successfully: {success_count}\n"
                    f"Objects failed: {failure_count}\n"
                    f"Skipped overwrite objects: {skipped_overwrite_count}\n\n"
                    f"Report saved to:\n{report_path}\n\n"
                    "Review the CSV report for per-object results."
                ),
                parent=self.root,
            )

    def _folder_copy_worker(self, source_folder_uri: str, destination_folder_uri: str) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            copy_items, preview = self._build_folder_copy_items(s3_client, source_folder_uri, destination_folder_uri)

            self._enqueue_ui(
                self._append_log,
                (
                    f"Folder scan complete. Found {preview.object_count} object(s). "
                    "Starting main copy pass."
                ),
            )
            self._execute_folder_copy_items(s3_client, copy_items, "folder_copy_result", "Folder Copy")
        except UserVisibleError as error:
            self._enqueue_ui(self._append_log, f"Folder copy failed: {error}")
            self._enqueue_ui(messagebox.showerror, "Folder Copy Failed", str(error), parent=self.root)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Folder Copy Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)

    def _bulk_folder_copy_worker(self, folder_jobs: list[tuple[str, str]]) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            copy_items, preview = self._build_bulk_folder_copy_items(s3_client, folder_jobs)

            self._enqueue_ui(
                self._append_log,
                (
                    f"Bulk folder scan complete. Found {preview.object_count} object(s) across "
                    f"{preview.folder_pair_count} folder pair(s). Starting main copy pass."
                ),
            )
            self._execute_folder_copy_items(s3_client, copy_items, "bulk_folder_copy_result", "Bulk Folder Copy")
        except UserVisibleError as error:
            self._enqueue_ui(self._append_log, f"Bulk folder copy failed: {error}")
            self._enqueue_ui(messagebox.showerror, "Bulk Folder Copy Failed", str(error), parent=self.root)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Bulk Folder Copy Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)

    def _inventory_worker(self, inventory_uri: str) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            bucket, prefix, normalized_uri = self._parse_s3_inventory_uri(inventory_uri)
            self._enqueue_ui(self._append_log, f"Starting inventory scan for {normalized_uri}")
            listed_objects = list_objects_with_metadata_under_prefix(
                s3_client,
                bucket,
                prefix,
                progress_callback=lambda msg: self._enqueue_ui(self._append_log, f"Inventory scan: {msg}"),
            )
            report_path = self._write_inventory_report(normalized_uri, listed_objects)
            self._enqueue_ui(
                self._append_log,
                f"Inventory export finished. Listed {len(listed_objects)} object(s). Report written to {report_path}",
            )
            self._enqueue_ui(
                messagebox.showinfo,
                "Inventory Export Complete",
                (
                    f"Objects listed: {len(listed_objects)}\n\n"
                    f"Report saved to:\n{report_path}"
                ),
                parent=self.root,
            )
        except UserVisibleError as error:
            self._enqueue_ui(self._append_log, f"Inventory export failed: {error}")
            self._enqueue_ui(messagebox.showerror, "Inventory Export Failed", str(error), parent=self.root)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Inventory Export Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)

    def _upload_worker(self, upload_items: list[DirectUploadItem]) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            for item in upload_items:
                self._upload_one_object(s3_client, item)

            self._enqueue_ui(self._append_log, "Upload succeeded. Local source file(s) were not modified.")
            self._play_completion_notification()
            success_lines = [
                f"Objects uploaded: {len(upload_items)}",
            ]
            if len(upload_items) <= 6:
                for item in upload_items:
                    success_lines.append(f"{item.label} Source: {item.local_path}")
                    success_lines.append(f"{item.label} Destination: {item.destination_uri}")
            else:
                success_lines.append("See Status / Output for per-item details.")
            self._enqueue_ui(
                messagebox.showinfo,
                "Success",
                (
                    "Direct upload completed successfully.\n\n"
                    + "\n".join(success_lines)
                ),
                parent=self.root,
            )
        except UserVisibleError as error:
            self._enqueue_ui(self._append_log, f"Direct upload failed: {error}")
            self._enqueue_ui(messagebox.showerror, "Direct Upload Failed", str(error), parent=self.root)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Direct Upload Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)

    def _rename_worker(
        self,
        source_ref: S3ObjectRef,
        destination_ref: S3ObjectRef,
        source_uri: str,
        dest_uri: str,
    ) -> None:
        try:
            credentials = self._get_active_credentials()
            s3_client = create_s3_client(self.config, credentials)
            self._rename_one_object(s3_client, source_ref, destination_ref, source_uri, dest_uri)

            self._enqueue_ui(self._append_log, "Rename succeeded. Old destination key was removed.")
            self._play_completion_notification()
            self._enqueue_ui(
                messagebox.showinfo,
                "Success",
                (
                    "Rename completed successfully.\n\n"
                    f"Old key: {source_uri}\n"
                    f"New key: {dest_uri}"
                ),
                parent=self.root,
            )
        except UserVisibleError as error:
            self._enqueue_ui(self._append_log, f"Rename failed: {error}")
            self._enqueue_ui(messagebox.showerror, "Rename Failed", str(error), parent=self.root)
        except RuntimeError as error:
            self._enqueue_ui(self._append_log, f"Configuration error: {error}")
            self._enqueue_ui(messagebox.showerror, "Configuration Error", str(error), parent=self.root)
        except Exception as error:  # pylint: disable=broad-except
            self._enqueue_ui(self._append_log, f"Unexpected failure: {error}")
            self._enqueue_ui(
                messagebox.showerror,
                "Rename Failed",
                f"Unexpected error: {error}",
                parent=self.root,
            )
        finally:
            self._enqueue_ui(self._set_running, False)


def main() -> None:
    root = tk.Tk()
    ttk.Style(root)
    app = S3CopyApp(root)
    _ = app
    root.mainloop()


if __name__ == "__main__":
    main()
